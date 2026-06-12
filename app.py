import streamlit as st
import pandas as pd
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import streamlit.components.v1 as components
import os

# Set page config
st.set_page_config(
    page_title="Salary Review Portal",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

DB_PATH = "salary_reviews.db"

# Custom CSS for Premium Glassmorphism and modern UI styling
st.markdown("""
<style>
    /* Global styles */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif;
    }
    
    /* Background adjustments */
    .stApp {
        background-color: #f8fafc;
    }
    
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #ffffff;
        border-right: 1px solid #e2e8f0;
        box-shadow: 2px 0 10px rgba(0, 0, 0, 0.02);
    }
    
    /* Premium KPI Card Styling */
    .kpi-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02), 0 2px 4px -1px rgba(0,0,0,0.02);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        margin-bottom: 15px;
    }
    .kpi-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 15px -3px rgba(0,0,0,0.05), 0 4px 6px -2px rgba(0,0,0,0.03);
    }
    .kpi-title {
        color: #64748b;
        font-size: 13px;
        font-weight: 600;
        letter-spacing: 0.05em;
        text-transform: uppercase;
        margin-bottom: 6px;
    }
    .kpi-value {
        color: #0f172a;
        font-size: 28px;
        font-weight: 700;
        line-height: 1;
    }
    .money-primary {
        color: #0f172a;
        font-weight: 700;
    }
    .money-secondary {
        color: #94a3b8;
        display: block;
        font-size: 1em;
        font-weight: 500;
        line-height: 1.2;
        margin-top: 2px;
    }
    .kpi-value .money-secondary {
        font-size: 0.58em;
    }
    .money-inline-secondary {
        color: #94a3b8;
        font-size: 0.86em;
        font-weight: 500;
        margin-left: 6px;
    }
    .kpi-sub {
        color: #10b981;
        font-size: 12px;
        margin-top: 6px;
        font-weight: 500;
    }
    .kpi-sub-neg {
        color: #f43f5e;
        font-size: 12px;
        margin-top: 6px;
        font-weight: 500;
    }
    .kpi-sub-neutral {
        color: #64748b;
        font-size: 12px;
        margin-top: 6px;
        font-weight: 500;
    }
    
    /* Section containers */
    .content-section {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 24px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.01);
    }
    
    /* Custom headers */
    .section-title {
        font-size: 20px;
        font-weight: 600;
        color: #0f172a;
        margin-bottom: 20px;
        border-bottom: 2px solid #f1f5f9;
        padding-bottom: 8px;
    }
    
    /* Button styles overrides */
    div.stButton > button {
        background: linear-gradient(135deg, #4f46e5 0%, #6366f1 100%);
        color: white;
        font-weight: 500;
        border: none;
        border-radius: 10px;
        padding: 10px 24px;
        box-shadow: 0 4px 6px rgba(99, 102, 241, 0.2);
        transition: all 0.2s ease;
    }
    div.stButton > button:hover {
        background: linear-gradient(135deg, #4338ca 0%, #4f46e5 100%);
        transform: translateY(-1px);
        box-shadow: 0 6px 12px rgba(99, 102, 241, 0.3);
        color: white;
    }
    div.stButton > button:active {
        transform: translateY(1px);
    }
    
    /* Tertiary buttons (flat icons) */
    div.stButton > button[kind="tertiary"] {
        background: transparent !important;
        color: #000000 !important;
        border: none !important;
        box-shadow: none !important;
        padding: 8px !important;
    }
    div.stButton > button[kind="tertiary"]:hover {
        background: #f1f5f9 !important;
        transform: none !important;
    }
    
    /* Table headers styling */
    th {
        background-color: #f8fafc !important;
        color: #475569 !important;
        font-weight: 600 !important;
    }
</style>
""", unsafe_allow_html=True)

# Database helper functions
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def execute_query(query, params=(), commit=False, hide_errors=False):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(query, params)
        if commit:
            conn.commit()
            result = cursor.rowcount
        else:
            result = [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        conn.rollback()
        if not hide_errors:
            st.error(f"Database error: {e}")
        result = None
    finally:
        conn.close()
    return result

def format_currency_inr(val):
    if val is None or pd.isna(val):
        return "N/A"
    return f"₹{float(val):,.0f}"

def format_currency(val, include_eur=True):
    if val is None or pd.isna(val):
        return "N/A"
    inr_amount = format_currency_inr(val)
    if not include_eur:
        return inr_amount
    eur_amount = format_eur(val, get_eur_rate())
    return f"{inr_amount} ({eur_amount})"

def format_currency_html(val, block=True):
    if val is None or pd.isna(val):
        return "N/A"
    inr_amount = format_currency_inr(val)
    eur_amount = format_eur(val, get_eur_rate())
    secondary_class = "money-secondary" if block else "money-inline-secondary"
    return (
        f"<span class='money-primary'>{inr_amount}</span>"
        f"<span class='{secondary_class}'>{eur_amount}</span>"
    )

def format_percentage(val):
    if val is None:
        return "0.0%"
    return f"{val:+.1f}%"

def format_display_date(date_str):
    if not date_str or pd.isna(date_str) or str(date_str).strip() in ("-", "nan", "None"):
        return "-"
    try:
        return datetime.strptime(str(date_str)[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return date_str

def init_settings():
    execute_query("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """, commit=True)
    existing = execute_query("SELECT value FROM settings WHERE key = 'eur_to_inr_rate'")
    if not existing:
        execute_query("INSERT INTO settings (key, value) VALUES (?, ?)", ('eur_to_inr_rate', '90.0'), commit=True)

init_settings()

def migrate_db():
    try:
        execute_query("ALTER TABLE employees ADD COLUMN last_review_date TEXT", commit=True, hide_errors=True)
        execute_query("ALTER TABLE employees ADD COLUMN last_review_effective_date TEXT", commit=True, hide_errors=True)
        execute_query("ALTER TABLE employees ADD COLUMN resign_date TEXT", commit=True, hide_errors=True)
        execute_query("ALTER TABLE employees ADD COLUMN lwd TEXT", commit=True, hide_errors=True)
        
        # Create archive table if not exists
        execute_query("""
            CREATE TABLE IF NOT EXISTS salary_reviews_archive (
                archive_id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_id INTEGER,
                employee_id INTEGER NOT NULL,
                review_name TEXT NOT NULL,
                review_date TEXT,
                previous_salary REAL,
                increment_amount REAL,
                increment_percentage REAL,
                new_salary REAL,
                effective_date TEXT,
                remark TEXT,
                status TEXT,
                archived_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
        """, commit=True, hide_errors=True)
        
        # Create bonuses table if not exists
        execute_query("""
            CREATE TABLE IF NOT EXISTS bonuses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                amount REAL,
                date_declared TEXT,
                due_date TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
        """, commit=True, hide_errors=True)

        # One-time backfill only — do not re-run on every app load (that overwrote manual DB/UI edits with NULL
        # when no Finalized salary_reviews row exists for the employee).
        already_backfilled = execute_query(
            "SELECT value FROM settings WHERE key = 'last_review_dates_backfilled'"
        )
        if not already_backfilled:
            execute_query("""
                UPDATE employees
                SET
                    last_review_date = COALESCE(
                        last_review_date,
                        (SELECT review_date FROM salary_reviews
                         WHERE employee_id = employees.id AND status = 'Finalized'
                         ORDER BY effective_date DESC, id DESC LIMIT 1)
                    ),
                    last_review_effective_date = COALESCE(
                        last_review_effective_date,
                        (SELECT effective_date FROM salary_reviews
                         WHERE employee_id = employees.id AND status = 'Finalized'
                         ORDER BY effective_date DESC, id DESC LIMIT 1)
                    )
            """, commit=True, hide_errors=True)
            execute_query(
                "INSERT INTO settings (key, value) VALUES (?, ?)",
                ('last_review_dates_backfilled', '1'),
                commit=True,
            )
    except Exception:
        # Columns might already exist
        pass

migrate_db()

def get_eur_rate():
    try:
        res = execute_query("SELECT value FROM settings WHERE key = 'eur_to_inr_rate'")
        if res:
            return float(res[0]['value'])
    except Exception:
        pass
    return 90.0

def format_eur(val, rate):
    if val is None or rate <= 0:
        return "N/A"
    return f"€{(val / rate):,.0f}"

def format_eur_rate_note(rate):
    if rate is None or rate <= 0:
        return "EUR conversion unavailable"
    return f"EUR conversion: 1 EUR = ₹{rate:,.2f}"

def calculate_tenure(doj_str):
    if not doj_str:
        return "N/A"
    try:
        joining_date = datetime.strptime(doj_str, "%Y-%m-%d").date()
        today = date.today()
        years = today.year - joining_date.year
        months = today.month - joining_date.month
        if months < 0:
            years -= 1
            months += 12
        
        parts = []
        if years > 0:
            parts.append(f"{years} yr" + ("s" if years > 1 else ""))
        if months > 0:
            parts.append(f"{months} mo" + ("s" if months > 1 else ""))
        
        return " ".join(parts) if parts else "Joined recently"
    except Exception:
        return doj_str

def format_tenure(tenure_val, doj_str):
    if tenure_val is not None and str(tenure_val).strip() != "" and not pd.isna(tenure_val):
        try:
            return f"{float(tenure_val):.1f} yrs"
        except ValueError:
            pass
    # Fallback to calculating on the fly as fractional years
    if not doj_str:
        return "N/A"
    try:
        doj = datetime.strptime(doj_str, "%Y-%m-%d").date()
        today = date.today()
        days_diff = (today - doj).days
        if days_diff < 0:
            return "0.0 yrs"
        val = round(days_diff / 365.25, 1)
        return f"{val:.1f} yrs"
    except Exception:
        return "N/A"

def num_to_words_indian(num):
    if num is None or pd.isna(num):
        return ""
    try:
        num = int(round(float(num)))
    except (ValueError, TypeError):
        return ""
    if num == 0:
        return "Zero Rupees"
    
    def helper(n):
        units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
                 "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        
        if n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        elif n < 1000:
            return units[n // 100] + " Hundred" + (" and " + helper(n % 100) if n % 100 != 0 else "")
        return ""

    parts = []
    
    # Crores (10,000,000+)
    if num >= 10000000:
        parts.append(helper(num // 10000000) + " Crore")
        num %= 10000000
        
    # Lakhs (100,000+)
    if num >= 100000:
        parts.append(helper(num // 100000) + " Lakh")
        num %= 100000
        
    # Thousands (1,000+)
    if num >= 1000:
        parts.append(helper(num // 1000) + " Thousand")
        num %= 1000
        
    # Hundreds / Tens / Units
    if num > 0:
        parts.append(helper(num))
        
    res = " ".join(parts)
    res = " ".join(res.split())
    return res.strip() + " Rupees"

def render_header(title, tag):
    col_title, col_tag = st.columns([8, 1])
    with col_title:
        st.markdown(f"<h1 style='margin-top: 0px; margin-bottom: 0px; color:#0f172a; font-weight:700;'>{title}</h1>", unsafe_allow_html=True)
    with col_tag:
        st.markdown(
            f"<div style='text-align: right; margin-top: 5px;'>"
            f"<span style='background-color:#ffe4e6; color:#b91c1c; border:1px solid #fecdd3; "
            f"padding:6px 18px; border-radius:10px; font-weight:700; font-size:15px; "
            f"display: inline-block; box-shadow: 0 2px 4px rgba(0,0,0,0.01);'>{tag}</span>"
            f"</div>",
            unsafe_allow_html=True
        )
    st.markdown("<div style='margin-top: 25px;'></div>", unsafe_allow_html=True)

@st.dialog("Confirm Deletion")
def confirm_delete_dialog(review_id):
    st.warning("Are you sure you want to delete this review? It will be moved to the archive.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Confirm Delete", type="primary", use_container_width=True):
            execute_query("INSERT INTO salary_reviews_archive (original_id, employee_id, review_name, review_date, previous_salary, increment_amount, increment_percentage, new_salary, effective_date, remark, status) SELECT id, employee_id, review_name, review_date, previous_salary, increment_amount, increment_percentage, new_salary, effective_date, remark, status FROM salary_reviews WHERE id = ?", (review_id,), commit=True)
            execute_query("DELETE FROM salary_reviews WHERE id = ?", (review_id,), commit=True)
            st.rerun()
    with c2:
        if st.button("Cancel", use_container_width=True):
            st.rerun()

import io

def to_excel(df):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        
        blue_font = Font(color="0070C0")
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            title_str = str(column_title)
            is_eur = "(EUR)" in title_str
            is_pct = "%" in title_str
            is_rev_sr = (title_str == "Rev Sr")
            
            header_cell = worksheet.cell(row=1, column=col_num)
            if '\n' in title_str:
                header_cell.alignment = Alignment(wrap_text=True)
                max_len = max([len(line) for line in title_str.split('\n')])
            else:
                max_len = len(title_str)
                
            col_letter = get_column_letter(col_num)
            
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                if is_eur:
                    cell.font = blue_font
                if is_pct:
                    cell.number_format = '0%'
                if is_rev_sr and cell.value == 1:
                    cell.fill = light_blue_fill
                    
                if cell.value is not None:
                    if is_pct and isinstance(cell.value, (int, float)):
                        val_len = len(str(int(cell.value * 100))) + 1
                    elif isinstance(cell.value, float):
                        val_len = len(f"{cell.value:.2f}")
                    else:
                        val_len = len(str(cell.value))
                    max_len = max(max_len, val_len)
            
            # Auto-adjust width with some padding, capped at 60
            worksheet.column_dimensions[col_letter].width = min(max_len + 3, 60)
                        
    processed_data = output.getvalue()
    return processed_data

def db_to_excel():
    conn = sqlite3.connect(DB_PATH)
    df_employees = pd.read_sql_query("SELECT * FROM employees", conn)
    df_reviews = pd.read_sql_query("SELECT * FROM salary_reviews", conn)
    try:
        df_bonuses = pd.read_sql_query("SELECT * FROM bonuses", conn)
    except Exception:
        df_bonuses = pd.DataFrame()
    conn.close()
    
    eur_rate = 90.0
    try:
        conn = sqlite3.connect(DB_PATH)
        res = conn.execute("SELECT value FROM settings WHERE key = 'eur_to_inr_rate'").fetchone()
        if res:
            eur_rate = float(res[0])
        conn.close()
    except Exception:
        pass

    # Add EUR columns in employees
    if not df_employees.empty and eur_rate > 0:
        df_employees.insert(df_employees.columns.get_loc('joining_salary') + 1, 'joining_salary_eur', 
                            df_employees['joining_salary'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))
        df_employees.insert(df_employees.columns.get_loc('current_salary') + 1, 'current_salary_eur', 
                            df_employees['current_salary'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))

    # Add EUR columns in reviews
    if not df_reviews.empty and eur_rate > 0:
        df_reviews.insert(df_reviews.columns.get_loc('previous_salary') + 1, 'previous_salary_eur', 
                           df_reviews['previous_salary'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))
        df_reviews.insert(df_reviews.columns.get_loc('increment_amount') + 1, 'increment_amount_eur', 
                           df_reviews['increment_amount'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))
        df_reviews.insert(df_reviews.columns.get_loc('new_salary') + 1, 'new_salary_eur', 
                           df_reviews['new_salary'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))

    # Add EUR columns in bonuses
    if not df_bonuses.empty and eur_rate > 0:
        df_bonuses.insert(df_bonuses.columns.get_loc('amount') + 1, 'amount_eur', 
                          df_bonuses['amount'].apply(lambda x: round(x / eur_rate, 2) if pd.notna(x) else None))

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_employees.to_excel(writer, index=False, sheet_name='Employees')
        df_reviews.to_excel(writer, index=False, sheet_name='Salary Reviews')
        if not df_bonuses.empty:
            df_bonuses.to_excel(writer, index=False, sheet_name='Bonuses')
    processed_data = output.getvalue()
    return processed_data

# Initialize DB check
if not os.path.exists(DB_PATH):
    st.warning("Database not initialized. Seeding database...")
    from db_init import init_db, import_excel
    init_db(force=True)
    import_excel()
    st.success("Seeding completed!")

# Sidebar navigation
st.sidebar.markdown(
    "<div style='padding: 10px 0; text-align: center;'>"
    "<h2 style='margin:0; color:#4f46e5; font-weight:700;'>Salary Review</h2>"
    "<p style='color:#64748b; font-size:12px; margin-top:2px;'>SQLite + Streamlit System</p>"
    "</div>", 
    unsafe_allow_html=True
)

st.sidebar.markdown("<hr style='margin: 10px 0; border: 0; border-top: 1px solid #e2e8f0;'/>", unsafe_allow_html=True)

# Sidebar Quick Stats
if "nav_menu" not in st.session_state:
    st.session_state.nav_menu = "📊 Dashboard Overview"

if "next_page" in st.session_state and st.session_state.next_page:
    st.session_state.nav_menu = st.session_state.next_page
    st.session_state.next_page = None

menu = st.sidebar.radio(
    "NAVIGATION",
    ["📊 Dashboard Overview", "📋 Employee Directory", "👤 Employee Profiles", "📜 Review History", "📈 Record New Review", "📅 Review Planner (Review-26)", "⚙️ Manage Employees", "🎁 Bonuses", "⚙️ System Settings"],
    key="nav_menu"
)

# Sidebar Quick Stats
active_count_query = execute_query("SELECT COUNT(*) as count FROM employees WHERE status = 'Active'")
total_payroll_query = execute_query("SELECT SUM(current_salary) as sum FROM employees WHERE status = 'Active'")
active_count = active_count_query[0]['count'] if active_count_query else 0
total_payroll = total_payroll_query[0]['sum'] if total_payroll_query else 0

st.sidebar.markdown("<div style='margin-top: 60px;'></div>", unsafe_allow_html=True)
st.sidebar.markdown(
    f"<div style='background-color:#f1f5f9; padding: 15px; border-radius:10px; border:1px solid #e2e8f0;'>"
    f"<div style='font-size:11px; color:#64748b; font-weight:600; text-transform:uppercase;'>System Status</div>"
    f"<div style='font-size:13px; font-weight:600; color:#0f172a; margin-top:5px;'>Active Headcount: <span style='color:#4f46e5;'>{active_count}</span></div>"
    f"<div style='font-size:13px; font-weight:600; color:#0f172a; margin-top:2px;'>Total Payroll: <span style='color:#10b981;'>{format_currency_html(total_payroll, block=False)}</span></div>"
    f"</div>",
    unsafe_allow_html=True
)

st.sidebar.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
db_excel_sidebar = db_to_excel()
st.sidebar.download_button(
    label="📥 Export Master DB (Excel)",
    data=db_excel_sidebar,
    file_name=f"salary_review_master_backup_{date.today().strftime('%Y-%m-%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)


# --- DASHBOARD PAGE ---
if menu == "📊 Dashboard Overview":
    render_header("Salary Review Dashboard", "S-1")
    eur_rate = get_eur_rate()
    st.caption(format_eur_rate_note(eur_rate))
    
    # Fetch KPI metrics
    resigned_query = execute_query("SELECT COUNT(*) as count FROM employees WHERE status = 'Resigned'")
    resigned_count = resigned_query[0]['count'] if resigned_query else 0
    
    avg_salary_query = execute_query("SELECT AVG(current_salary) as avg FROM employees WHERE status = 'Active'")
    avg_salary = avg_salary_query[0]['avg'] if avg_salary_query else 0
    
    proposed_reviews_query = execute_query("SELECT COUNT(*) as count, SUM(increment_amount) as total_inc, AVG(increment_percentage) as avg_pct FROM salary_reviews WHERE status = 'Proposed'")
    proposed_count = proposed_reviews_query[0]['count'] if proposed_reviews_query else 0
    proposed_total_inc = proposed_reviews_query[0]['total_inc'] if proposed_reviews_query else 0
    proposed_avg_pct = proposed_reviews_query[0]['avg_pct'] if proposed_reviews_query else 0
    
    # 4 Columns for KPIs
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(
            f"<div class='kpi-card'>"
            f"<div class='kpi-title'>Active Headcount</div>"
            f"<div class='kpi-value'><span class='money-primary'>{active_count}</span><span class='money-secondary'>&nbsp;</span></div>"
            f"<div class='kpi-sub-neutral'>{resigned_count} Resigned / Inactive</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        
    with col2:
        st.markdown(
            f"<div class='kpi-card'>"
            f"<div class='kpi-title'>Active Payroll</div>"
            f"<div class='kpi-value'>{format_currency_html(total_payroll)}</div>"
            f"<div class='kpi-sub'>Annualized base pay</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        
    with col3:
        st.markdown(
            f"<div class='kpi-card'>"
            f"<div class='kpi-title'>Average Salary</div>"
            f"<div class='kpi-value'>{format_currency_html(avg_salary)}</div>"
            f"<div class='kpi-sub-neutral'>Per active resource</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        
    with col4:
        sub_text_style = "kpi-sub" if (proposed_avg_pct or 0) > 0 else "kpi-sub-neutral"
        st.markdown(
            f"<div class='kpi-card'>"
            f"<div class='kpi-title'>Proposed Reviews</div>"
            f"<div class='kpi-value'>{format_currency_html(proposed_total_inc or 0)}</div>"
            f"<div class='{sub_text_style}'>+{proposed_avg_pct or 0:.2f}% avg increment ({proposed_count} employees)</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        
    # Main Content Area - Full Width
    st.markdown("<div class='content-section'>", unsafe_allow_html=True)
    c_title, c_btn, c_exp = st.columns([2.2, 1.0, 0.8])
    with c_title:
        st.markdown("<div class='section-title'>Active Employees Salary Sheet</div>", unsafe_allow_html=True)
    with c_btn:
        if st.button("🔄 Update Tenure", type="secondary", use_container_width=True, help="Recalculate tenure for all employees and save to database"):
            emps = execute_query("SELECT id, date_of_joining FROM employees")
            updated_count = 0
            for emp in emps:
                doj_str = emp['date_of_joining']
                if doj_str:
                    try:
                        doj = datetime.strptime(doj_str, "%Y-%m-%d").date()
                        today = date.today()
                        days_diff = (today - doj).days
                        tenure_val = round(days_diff / 365.25, 1) if days_diff >= 0 else 0.0
                    except Exception:
                        tenure_val = None
                else:
                    tenure_val = None
                
                execute_query("UPDATE employees SET tenure = ? WHERE id = ?", (tenure_val, emp['id']), commit=True)
                updated_count += 1
            st.success(f"Updated tenure for {updated_count} employees!")
            st.rerun()
            
    # Load employees data
    employees = execute_query("""
        SELECT id, name, date_of_joining, role, joining_salary, current_salary, status, tenure 
        FROM employees 
        ORDER BY current_salary DESC
    """)
    
    if employees:
        df_emp = pd.DataFrame(employees)
        df_emp['Tenure'] = df_emp.apply(lambda r: format_tenure(r.get('tenure'), r['date_of_joining']), axis=1)

        # Build proposed data lookup
        proposed_map = {}
        for emp_id in df_emp['id']:
            prop = execute_query(
                "SELECT increment_amount, increment_percentage, new_salary FROM salary_reviews WHERE employee_id = ? AND status = 'Proposed'",
                (emp_id,)
            )
            if prop:
                proposed_map[emp_id] = prop[0]

        has_proposed = len(proposed_map) > 0
        
        with c_exp:
            df_export = df_emp.copy()
            df_export['Emp ID']          = df_export['id']
            df_export['Role']            = df_export['role'].fillna('-')
            df_export['Date of Joining'] = df_export['date_of_joining'].apply(format_display_date)
            
            eur_rate = get_eur_rate()
            df_export['Joining Salary (INR)'] = df_export['joining_salary'].apply(format_currency_inr)
            df_export['Joining Salary (EUR)'] = df_export['joining_salary'].apply(lambda x: format_eur(x, eur_rate))
            df_export['Current Salary (INR)'] = df_export['current_salary'].apply(format_currency_inr)
            df_export['Current Salary (EUR)'] = df_export['current_salary'].apply(lambda x: format_eur(x, eur_rate))
            df_export['Status']          = df_export['status']
            
            export_cols = ['Emp ID', 'name', 'Role', 'Date of Joining', 'Tenure', 
                           'Joining Salary (INR)', 'Joining Salary (EUR)', 
                           'Current Salary (INR)', 'Current Salary (EUR)', 'Status']
            
            if has_proposed:
                df_export['Proposed Increment (INR)'] = df_export['id'].apply(lambda i: format_currency_inr(proposed_map[i]['increment_amount']) if i in proposed_map else '-')
                df_export['Proposed Increment (EUR)'] = df_export['id'].apply(lambda i: format_eur(proposed_map[i]['increment_amount'], eur_rate) if i in proposed_map and pd.notna(proposed_map[i]['increment_amount']) else '-')
                df_export['Proposed %']         = df_export['id'].apply(lambda i: format_percentage(proposed_map[i]['increment_percentage']) if i in proposed_map else '-')
                df_export['New Salary (Proj) (INR)']  = df_export['id'].apply(lambda i: format_currency_inr(proposed_map[i]['new_salary']) if i in proposed_map else '-')
                df_export['New Salary (Proj) (EUR)']  = df_export['id'].apply(lambda i: format_eur(proposed_map[i]['new_salary'], eur_rate) if i in proposed_map and pd.notna(proposed_map[i]['new_salary']) else '-')
                export_cols += ['Proposed Increment (INR)', 'Proposed Increment (EUR)', 'Proposed %', 'New Salary (Proj) (INR)', 'New Salary (Proj) (EUR)']
                
            df_final = df_export[export_cols].rename(columns={'name': 'Employee Name'})
            st.download_button(
                label="📥 Export to XLSX",
                data=to_excel(df_final),
                file_name=f"active_employees_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # --- S-7-style header row ---
        if has_proposed:
            hc0, hc1, hc2, hc3, hc4, hc5, hc6, hc7, hc8 = st.columns([0.6, 1.8, 1.5, 1.5, 1.5, 0.8, 1.7, 1.2, 1.7])
        else:
            hc0, hc1, hc2, hc3, hc4, hc5 = st.columns([0.6, 2.2, 1.6, 2.2, 2.2, 1.2])

        with hc0: st.markdown("**Emp ID**")
        with hc1: st.markdown("**Employee Name**")
        with hc2: st.markdown("**DoJ / Tenure**")
        with hc3: st.markdown("**Joining<br>Salary**", unsafe_allow_html=True)
        with hc4: st.markdown("**Current<br>Salary**", unsafe_allow_html=True)
        with hc5: st.markdown("**Status**")
        if has_proposed:
            with hc6: st.markdown("**Proposed Increment**")
            with hc7: st.markdown("**Proposed %**")
            with hc8: st.markdown("**New<br>Salary**", unsafe_allow_html=True)
        st.markdown("<hr style='margin: 0.35rem 0 0.75rem 0;'>", unsafe_allow_html=True)

        # --- Rows ---
        for _, row in df_emp.iterrows():
            prop = proposed_map.get(row['id'])
            if has_proposed:
                c0, c1, c2, c3, c4, c5, c6, c7, c8 = st.columns([0.6, 1.8, 1.5, 1.5, 1.5, 0.8, 1.7, 1.2, 1.7])
            else:
                c0, c1, c2, c3, c4, c5 = st.columns([0.6, 2.2, 1.6, 2.2, 2.2, 1.2])

            with c0: st.write(row['id'])
            with c1:
                name_val = row['name']
                role_val = row.get('role') or 'N/A'
                st.markdown(f"<div style='font-weight:600;'>{name_val}</div><div style='font-size:0.85em; color:#64748b;'>{role_val}</div>", unsafe_allow_html=True)
            with c2:
                s_date = format_display_date(row['date_of_joining'])
                t_str = row['Tenure']
                st.markdown(f"<div style='font-size:0.9em;'>{s_date}<br><span style='color:#64748b;'>{t_str}</span></div>", unsafe_allow_html=True)
            with c3: st.markdown(format_currency_html(row['joining_salary'], block=True), unsafe_allow_html=True)
            with c4: st.markdown(format_currency_html(row['current_salary'], block=True), unsafe_allow_html=True)
            with c5:
                st.write(row['status'])
            if has_proposed:
                with c6:
                    if prop: st.markdown(format_currency_html(prop['increment_amount'], block=True), unsafe_allow_html=True)
                    else: st.write("-")
                with c7: st.write(format_percentage(prop['increment_percentage']) if prop else "-")
                with c8:
                    if prop: st.markdown(format_currency_html(prop['new_salary'], block=True), unsafe_allow_html=True)
                    else: st.write("-")

        st.markdown("<div style='margin-top:15px;'></div>", unsafe_allow_html=True)


    else:
        st.info("No employees found in the database.")
    st.markdown("</div>", unsafe_allow_html=True)
        



# --- EMPLOYEE DIRECTORY PAGE ---
elif menu == "📋 Employee Directory":
    render_header("Employee Directory", "S-2")
    
    if True:
        # Load all employees
        employees = execute_query("""
            SELECT id, name, date_of_joining, role, current_salary, status, department, resign_date, tenure 
            FROM employees 
            ORDER BY name ASC
        """)
        
        if employees:
            df_emp = pd.DataFrame(employees)
            df_emp['Tenure'] = df_emp.apply(lambda r: format_tenure(r.get('tenure'), r['date_of_joining']), axis=1)
            df_emp['status'] = df_emp.apply(lambda r: 'Resigned' if pd.notna(r.get('resign_date')) and str(r.get('resign_date')).strip() else r['status'], axis=1)
            
            # Search & Filter bar
            col_s1, col_s2, col_s3 = st.columns([2, 1, 1])
            with col_s1:
                search_query = st.text_input("Search by Name:", value="", placeholder="Type name to search...")
            with col_s2:
                status_filter = st.selectbox("Status:", ["All Statuses", "Active", "Resigned"])
            with col_s3:
                sort_by = st.selectbox("Sort By:", ["Name (A-Z)", "Salary (High-Low)", "Date of Joining (Oldest-Newest)"])
                
            # Apply filters
            df_filtered = df_emp.copy()
            if search_query:
                df_filtered = df_filtered[df_filtered['name'].str.contains(search_query, case=False, na=False)]
            if status_filter != "All Statuses":
                df_filtered = df_filtered[df_filtered['status'] == status_filter]
                
            # Apply sorting
            if sort_by == "Name (A-Z)":
                df_filtered = df_filtered.sort_values(by='name')
            elif sort_by == "Salary (High-Low)":
                df_filtered = df_filtered.sort_values(by='current_salary', ascending=False)
            elif sort_by == "Date of Joining (Oldest-Newest)":
                df_filtered = df_filtered.sort_values(by='date_of_joining')
                
            # Stats below filter
            col_stat1, col_stat2 = st.columns([3, 1])
            with col_stat1:
                st.write(f"Showing **{len(df_filtered)}** employees matching criteria.")
            with col_stat2:
                # generate export and place button
                df_table = df_filtered.copy()
                df_table['Joining Date (DoJ)'] = df_table['date_of_joining'].apply(format_display_date)
                
                eur_rate = get_eur_rate()
                df_table['Current Salary (INR)'] = df_table['current_salary'].apply(format_currency_inr)
                df_table['Current Salary (EUR)'] = df_table['current_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_table['Role'] = df_table['role'].fillna('-')
                df_table['Department'] = df_table['department'].fillna('-')
                
                df_table_show = df_table[['id', 'name', 'Role', 'Department', 'Joining Date (DoJ)', 'Tenure', 'Current Salary (INR)', 'Current Salary (EUR)', 'status']].rename(columns={'id': 'Emp ID', 'name': 'Employee Name', 'status': 'Status'})
                st.download_button(label="📥 Export to XLSX", data=to_excel(df_table_show), file_name=f"employee_directory_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            
            # Display Tab views: Table View vs Card View
            tab_view1, tab_view2 = st.tabs(["📋 Table View", "🗂️ Card View"])
            
            with tab_view1:
                st.markdown("<div class='content-section'>", unsafe_allow_html=True)
                
                # Table Header
                hcol_sr, hcol0, hcol1, hcol_role, hcol2, hcol3, hcol4, hcol5 = st.columns([0.5, 0.8, 2.0, 1.5, 1.5, 1.2, 2.0, 1.0])
                with hcol_sr: st.markdown("**Sr No**")
                with hcol0: st.markdown("**Emp ID**")
                with hcol1: st.markdown("**Employee Name**")
                with hcol_role: st.markdown("**Role**")
                with hcol2: st.markdown("**DoJ**")
                with hcol3: st.markdown("**Tenure**")
                with hcol4: st.markdown("**Current<br>Salary**", unsafe_allow_html=True)
                with hcol5: st.markdown("**Status**")
                st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
                
                virtual_sr = 1
                for idx, row in df_filtered.iterrows():
                    col_sr, col0, col1, col_role, col2, col3, col4, col5 = st.columns([0.5, 0.8, 2.0, 1.5, 1.5, 1.2, 2.0, 1.0])
                    with col_sr: st.write(virtual_sr)
                    with col0: st.write(row['id'])
                    with col1: st.write(row['name'])
                    with col_role: st.write(row.get('role') or 'N/A')
                    with col2: st.write(format_display_date(row['date_of_joining']))
                    with col3: st.write(row['Tenure'])
                    with col4: st.markdown(format_currency_html(row['current_salary'], block=True), unsafe_allow_html=True)
                    with col5: st.write(row['status'])
                    virtual_sr += 1
                
                st.markdown("<div style='margin-top:15px;'></div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
                
            with tab_view2:
                # Render grid cards
                card_cols = st.columns(3)
                for idx, row in enumerate(df_filtered.to_dict('records')):
                    col_idx = idx % 3
                    with card_cols[col_idx]:
                        st.markdown(
                            f"<div class='kpi-card' style='position: relative; margin-bottom: 5px;'>"
                            f"<div style='font-size:18px; font-weight:700; color:#0f172a;'>{row['name']}</div>"
                            f"<div style='font-size:12px; font-weight:500; color:#64748b; margin-top:2px;'>ID: {row['id']} | Role: {row.get('role') or 'N/A'} | Department: {row['department'] or 'Tech'}</div>"
                            f"<hr style='margin:10px 0; border:0; border-top:1px solid #f1f5f9;'/>"
                            f"<div style='font-size:13px; color:#475569;'><b>DoJ:</b> {format_display_date(row['date_of_joining'])}</div>"
                            f"<div style='font-size:13px; color:#475569;'><b>Tenure:</b> {row['Tenure']}</div>"
                            f"<div style='font-size:13px; color:#475569; margin-top:4px;'><b>Salary:</b> {format_currency_html(row['current_salary'], block=False)}</div>"
                            f"<div style='margin-top:12px; text-align:right;'>"
                            f"<span style='background-color:{'#d1fae5' if row['status'] == 'Active' else '#ffe4e6'}; color:{'#065f46' if row['status'] == 'Active' else '#991b1b'}; padding:4px 8px; border-radius:12px; font-size:11px; font-weight:600;'>{row['status']}</span>"
                            f"</div>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                        st.markdown("<div style='margin-bottom:20px;'></div>", unsafe_allow_html=True)
        else:
            st.info("No employee records found in the database.")


# --- REVIEW HISTORY PAGE ---
elif menu == "📜 Review History":
    if 'history_edit_id' not in st.session_state:
        st.session_state.history_edit_id = None

    if st.session_state.history_edit_id is not None:
        render_header("Edit Review Record", "S-4-Edit")
        components.html("<script>window.parent.scrollTo(0, 0);</script>", height=0)
        review_id = st.session_state.history_edit_id
        
        rev_rows = execute_query("""
            SELECT r.id, e.name as emp_name, r.review_name, r.review_date, r.previous_salary, 
                   r.increment_amount, r.increment_percentage, r.new_salary, r.effective_date, 
                   r.remark, r.status, r.sr_no
            FROM salary_reviews r
            JOIN employees e ON r.employee_id = e.id
            WHERE r.id = ?
        """, (review_id,))
        
        if not rev_rows:
            st.warning("Review record not found.")
            if st.button("← Back to Review History"):
                st.session_state.history_edit_id = None
                st.rerun()
        else:
            rev_data = rev_rows[0]
            if st.button("← Back to Review History"):
                st.session_state.history_edit_id = None
                st.rerun()
                
            st.markdown("<div class='content-section'>", unsafe_allow_html=True)
            st.markdown(f"<div class='section-title'>Edit Review: {rev_data['emp_name']}</div>", unsafe_allow_html=True)
            
            with st.form("edit_history_review_form"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    edit_rev_name = st.text_input("Review Label:", value=rev_data['review_name'] or "")
                with c2:
                    edit_rev_date = st.date_input("Review Date:", value=datetime.strptime(rev_data['review_date'], "%Y-%m-%d").date() if rev_data['review_date'] else date.today(), format="DD-MM-YYYY")
                with c3:
                    edit_eff_date = st.date_input("Effective Date:", value=datetime.strptime(rev_data['effective_date'], "%Y-%m-%d").date() if rev_data['effective_date'] else date.today(), format="DD-MM-YYYY")
                    
                c4, c5, c6 = st.columns(3)
                with c4:
                    edit_prev_sal = st.number_input("Previous Salary (₹):", min_value=0.0, value=float(rev_data['previous_salary'] or 0), step=5000.0)
                    if edit_prev_sal > 0:
                        st.caption(f"💶 **EUR:** {format_eur(edit_prev_sal, get_eur_rate())} | ✍️ {num_to_words_indian(edit_prev_sal)}")
                with c5:
                    edit_inc_amt = st.number_input("Increment Amount (₹):", value=float(rev_data['increment_amount'] or 0), step=1000.0)
                    if edit_inc_amt > 0:
                        st.caption(f"💶 **EUR:** {format_eur(edit_inc_amt, get_eur_rate())} | ✍️ {num_to_words_indian(edit_inc_amt)}")
                with c6:
                    calc_new_sal = edit_prev_sal + edit_inc_amt
                    st.number_input("New Salary (₹) - auto calculated", value=float(calc_new_sal), disabled=True)
                    if calc_new_sal > 0:
                        st.caption(f"💶 **EUR:** {format_eur(calc_new_sal, get_eur_rate())} | ✍️ {num_to_words_indian(calc_new_sal)}")
                    
                c7, c8, c9 = st.columns(3)
                with c7:
                    edit_sr_no = st.number_input("Sr No:", value=float(rev_data['sr_no']) if rev_data['sr_no'] is not None else None, step=1.0)
                with c8:
                    edit_status = st.selectbox("Status:", ["Proposed", "Finalized"], index=0 if rev_data['status'] == 'Proposed' else 1)
                with c9:
                    st.write("")
                    
                edit_rem = st.text_area("Remark:", value=rev_data['remark'] or "")
                
                submit_edit = st.form_submit_button("💾 Save Review Record", type="primary")
                if submit_edit:
                    new_sal = edit_prev_sal + edit_inc_amt
                    pct = (edit_inc_amt / edit_prev_sal * 100.0) if edit_prev_sal > 0 else 0.0
                    sr_no_val = int(edit_sr_no) if edit_sr_no is not None else None
                    execute_query("""
                        UPDATE salary_reviews
                        SET review_name = ?, review_date = ?, previous_salary = ?, increment_amount = ?, increment_percentage = ?, new_salary = ?, effective_date = ?, remark = ?, status = ?, sr_no = ?
                        WHERE id = ?
                    """, (edit_rev_name, edit_rev_date.strftime("%Y-%m-%d"), edit_prev_sal, edit_inc_amt, pct, new_sal, edit_eff_date.strftime("%Y-%m-%d"), edit_rem, edit_status, sr_no_val, review_id), commit=True)
                    
                    st.success("Review record updated successfully!")
                    st.session_state.history_edit_id = None
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
            st.stop()
            
    # --- END EDIT MODE ---

    render_header("Salary Review History Log", "S-4")
    # Load all reviews — include review id for edit button
    reviews_all = execute_query("""
        SELECT r.id as review_id, e.id as emp_id, e.name, e.date_of_joining, e.role, r.review_name, r.review_date, r.previous_salary, r.increment_amount, r.increment_percentage, r.new_salary, r.effective_date, r.remark, r.status, r.sr_no
        FROM salary_reviews r
        JOIN employees e ON r.employee_id = e.id
        ORDER BY r.effective_date DESC, r.id DESC
    """)
    
    if reviews_all:
        df_revs = pd.DataFrame(reviews_all)
        
        # Load unique review names for dropdown filter
        unique_labels = sorted(list(df_revs['review_name'].dropna().unique()))
        
        # Search & Filter bar
        col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns([1.5, 1, 1, 1.2, 1])
        with col_f1:
            unique_emp_names = sorted(list(df_revs['name'].dropna().unique()))
            search_name = st.selectbox("Search Employee Name:", ["All Employees"] + unique_emp_names)
        with col_f2:
            status_filter = st.selectbox("Review Status:", ["All Reviews", "Proposed", "Others"])
        with col_f3:
            label_filter = st.selectbox("Review Label:", ["All Labels"] + unique_labels)
        with col_f4:
            sort_field = st.selectbox("Sort Field:", ["Employee Name", "Effective Date", "Increment Amount", "Increment Percentage"])
        with col_f5:
            sort_order = st.selectbox("Sort Order:", ["Ascending", "Descending"])
            
        # Date Filters
        st.markdown("<div style='margin-bottom: 10px;'></div>", unsafe_allow_html=True)
        col_d1, col_d2, col_d3, col_d4 = st.columns([1, 1.5, 1, 1.5])
        with col_d1:
            eff_date_preset = st.selectbox("Effective Date:", ["All Time", "This Month", "Last Month", "This Year", "Last Year", "Custom Range"], key="eff_date_preset")
        with col_d2:
            eff_date_range = []
            if eff_date_preset == "Custom Range":
                eff_date_range = st.date_input("Effective Date Range:", value=[], key="eff_date_range")
            else:
                st.write("") # placeholder
                
        with col_d3:
            rev_date_preset = st.selectbox("Revision Date:", ["All Time", "This Month", "Last Month", "This Year", "Last Year", "Custom Range"], key="rev_date_preset")
        with col_d4:
            rev_date_range = []
            if rev_date_preset == "Custom Range":
                rev_date_range = st.date_input("Revision Date Range:", value=[], key="rev_date_range")
            else:
                st.write("")

        def get_date_bounds(preset, custom_range):
            import calendar
            from datetime import date, timedelta
            today = date.today()
            if preset == "All Time": return None, None
            elif preset == "This Month":
                return date(today.year, today.month, 1), date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
            elif preset == "Last Month":
                first_of_this = date(today.year, today.month, 1)
                last_of_prev = first_of_this - timedelta(days=1)
                return date(last_of_prev.year, last_of_prev.month, 1), last_of_prev
            elif preset == "This Year":
                return date(today.year, 1, 1), date(today.year, 12, 31)
            elif preset == "Last Year":
                return date(today.year - 1, 1, 1), date(today.year - 1, 12, 31)
            elif preset == "Custom Range":
                if custom_range and len(custom_range) == 2:
                    return custom_range[0], custom_range[1]
                elif custom_range and len(custom_range) == 1:
                    return custom_range[0], custom_range[0]
            return None, None
            
        # Apply filters
        df_filtered = df_revs.copy()
        if search_name != "All Employees":
            df_filtered = df_filtered[df_filtered['name'] == search_name]
            
        if status_filter == "Proposed":
            df_filtered = df_filtered[df_filtered['status'] == "Proposed"]
        elif status_filter == "Others":
            df_filtered = df_filtered[df_filtered['status'] != "Proposed"]
            
        if label_filter != "All Labels":
            df_filtered = df_filtered[df_filtered['review_name'] == label_filter]
            
        eff_start, eff_end = get_date_bounds(eff_date_preset, eff_date_range)
        if eff_start and eff_end:
            df_filtered = df_filtered[df_filtered['effective_date'].notna()]
            df_filtered = df_filtered[(df_filtered['effective_date'] >= eff_start.strftime("%Y-%m-%d")) & (df_filtered['effective_date'] <= eff_end.strftime("%Y-%m-%d"))]
            
        rev_start, rev_end = get_date_bounds(rev_date_preset, rev_date_range)
        if rev_start and rev_end:
            df_filtered = df_filtered[df_filtered['review_date'].notna()]
            df_filtered = df_filtered[(df_filtered['review_date'] >= rev_start.strftime("%Y-%m-%d")) & (df_filtered['review_date'] <= rev_end.strftime("%Y-%m-%d"))]
            
        # Apply sorting
        asc = (sort_order == "Ascending")
        if sort_field == "Effective Date":
            df_filtered = df_filtered.sort_values(by=['effective_date', 'sr_no'], ascending=[asc, True])
        elif sort_field == "Increment Amount":
            df_filtered = df_filtered.sort_values(by=['increment_amount', 'sr_no'], ascending=[asc, True])
        elif sort_field == "Increment Percentage":
            df_filtered = df_filtered.sort_values(by=['increment_percentage', 'sr_no'], ascending=[asc, True])
        elif sort_field == "Employee Name":
            df_filtered = df_filtered.sort_values(by=['name', 'sr_no'], ascending=[asc, True])

        if 'import_msg' in st.session_state:
            st.success(st.session_state.import_msg)
            del st.session_state.import_msg

        col_stat1, col_stat2, col_stat3, col_stat4, col_stat5 = st.columns([1.4, 1.1, 1.1, 1.1, 1.1])
        with col_stat1:
            st.write(f"Found **{len(df_filtered)}** review logs matching criteria.")
        with col_stat2:
            if st.button("➕ Add New Review", key="s4_add_new_btn", type="primary", use_container_width=True):
                st.session_state.next_page = "📈 Record New Review"
                st.rerun()
        with col_stat3:
            active_emps = execute_query("SELECT id as 'Emp ID', name as 'Employee Name', role as 'Role', date_of_joining as 'Date of Joining', current_salary as 'Previous Salary' FROM employees WHERE status='Active'")
            if active_emps:
                eur_rate = get_eur_rate()
                df_tmpl = pd.DataFrame(active_emps)
                df_tmpl.insert(0, 'Sr No', '')
                df_tmpl['Date of Joining'] = df_tmpl['Date of Joining'].apply(format_display_date)
                df_tmpl['Role'] = df_tmpl['Role'].fillna('-')
                
                df_tmpl['Previous Salary (INR)'] = df_tmpl['Previous Salary'].apply(format_currency_inr)
                df_tmpl['Previous Salary (EUR)'] = df_tmpl['Previous Salary'].apply(lambda x: format_eur(x, eur_rate))
                
                df_tmpl['Review Label'] = ''
                df_tmpl['Review Date'] = ''
                df_tmpl['Effective Date'] = ''
                df_tmpl['Increment Amount (INR)'] = ''
                df_tmpl['Increment Amount (EUR)'] = ''
                df_tmpl['New Salary (INR)'] = ''
                df_tmpl['New Salary (EUR)'] = ''
                df_tmpl['Status'] = 'Proposed'
                df_tmpl['Remark'] = ''
                
                df_tmpl = df_tmpl[['Sr No', 'Emp ID', 'Employee Name', 'Role', 'Date of Joining', 'Review Label', 'Review Date', 'Effective Date', 
                                   'Previous Salary (INR)', 'Previous Salary (EUR)', 'Increment Amount (INR)', 'Increment Amount (EUR)', 
                                   'New Salary (INR)', 'New Salary (EUR)', 'Status', 'Remark']]
                st.download_button("📥 Export Template", data=to_excel(df_tmpl), file_name="review_import_template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                
        with col_stat4:
            df_export = df_filtered.copy()
            df_export = df_export.sort_values(by=['name', 'sr_no'])
            eur_rate = get_eur_rate()
            
            df_export['Review Date']     = df_export['review_date'].apply(format_display_date)
            df_export['Effective Date']  = df_export['effective_date'].apply(format_display_date)
            
            df_export['Sr No']           = range(1, len(df_export) + 1)
            df_export['Rev Sr']          = df_export['sr_no']
            
            df_export['Role']            = df_export['role'].fillna('-')
            df_export['Date of Joining'] = df_export['date_of_joining'].apply(format_display_date)
            
            df_export['Previous Salary\n(INR)'] = df_export['previous_salary']
            df_export['Previous Salary\n(EUR)'] = df_export['previous_salary'].apply(lambda x: int(round(x / eur_rate)) if eur_rate > 0 and pd.notna(x) else None).astype('Int64')
            
            df_export['Increment\n(INR)'] = df_export['increment_amount']
            df_export['Increment\n(EUR)'] = df_export['increment_amount'].apply(lambda x: int(round(x / eur_rate)) if eur_rate > 0 and pd.notna(x) else None).astype('Int64')
            
            df_export['Increment %']     = df_export['increment_percentage'].apply(lambda x: float(round(x)) / 100.0 if pd.notna(x) else None)
            
            df_export['New Salary\n(INR)'] = df_export['new_salary']
            df_export['New Salary\n(EUR)'] = df_export['new_salary'].apply(lambda x: int(round(x / eur_rate)) if eur_rate > 0 and pd.notna(x) else None).astype('Int64')
            
            df_export['Status']          = df_export['status']
            df_export['Remark']          = df_export['remark'].fillna('-')
            
            export_cols = ['Sr No', 'emp_id', 'Rev Sr', 'name', 'Role', 'Date of Joining', 'review_name', 'Review Date', 'Effective Date',
                           'Previous Salary\n(INR)', 'Previous Salary\n(EUR)', 'Increment\n(INR)', 'Increment\n(EUR)',
                           'Increment %', 'New Salary\n(INR)', 'New Salary\n(EUR)', 'Status', 'Remark']
                           
            export_headers = {
                'emp_id': 'Emp ID', 'name': 'Employee Name', 'review_name': 'Review Label'
            }
            
            df_table_show = df_export[export_cols].rename(columns=export_headers)
            st.download_button(label="📥 Export to XLSX", data=to_excel(df_table_show), file_name=f"salary_reviews_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with col_stat5:
            with st.popover("⬆️ Import", use_container_width=True):
                uploaded_file = st.file_uploader("Upload filled template (XLSX format)", type=["xlsx"])
                if uploaded_file is not None:
                    if st.button("Process Import", type="primary"):
                        try:
                            df_import = pd.read_excel(uploaded_file)
                            success_count = 0
                            def parse_import_date(d):
                                if pd.isna(d): return None
                                try:
                                    return pd.to_datetime(d).strftime("%Y-%m-%d")
                                except:
                                    return None
                                    
                            for _, row in df_import.iterrows():
                                if 'Emp ID' not in row or pd.isna(row['Emp ID']): continue
                                emp_id = int(row['Emp ID'])
                                
                                emp = execute_query("SELECT id FROM employees WHERE id=?", (emp_id,))
                                if not emp: continue
                                
                                r_date = parse_import_date(row.get('Review Date'))
                                e_date = parse_import_date(row.get('Effective Date'))
                                
                                # Helper to get from multiple potential columns
                                def get_col_val(r, cols, default=0.0):
                                    for c in cols:
                                        if c in r and pd.notna(r[c]):
                                            try: return float(r[c])
                                            except: pass
                                    return default

                                prev_sal = get_col_val(row, ['Previous Salary', 'Previous Salary (INR)', 'Previous Salary\n(INR)', 'Previous Salary (EUR)', 'Previous Salary\n(EUR)'], 0.0)
                                inc_amt = get_col_val(row, ['Increment Amount', 'Increment Amount (INR)', 'Increment Amount\n(INR)', 'Increment (INR)', 'Increment\n(INR)', 'Increment Amount (EUR)', 'Increment Amount\n(EUR)', 'Increment (EUR)', 'Increment\n(EUR)'], 0.0)
                                new_sal = get_col_val(row, ['New Salary', 'New Salary (INR)', 'New Salary\n(INR)', 'New Salary (EUR)', 'New Salary\n(EUR)'], prev_sal + inc_amt)
                                
                                sr_no = int(row.get('Sr No')) if 'Sr No' in row and pd.notna(row.get('Sr No')) else None
                                label = str(row.get('Review Label', '')) if pd.notna(row.get('Review Label')) else ''
                                status = str(row.get('Status', 'Proposed')) if pd.notna(row.get('Status')) else 'Proposed'
                                remark = str(row.get('Remark', '')) if pd.notna(row.get('Remark')) else ''
                                
                                pct = (inc_amt / prev_sal * 100.0) if prev_sal > 0 else 0.0
                                
                                execute_query("""
                                    INSERT INTO salary_reviews (employee_id, review_name, review_date, previous_salary, increment_amount, increment_percentage, new_salary, effective_date, status, remark, sr_no)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (emp_id, label, r_date, prev_sal, inc_amt, pct, new_sal, e_date, status, remark, sr_no), commit=True)
                                success_count += 1
                                
                            st.session_state.import_msg = f"Successfully imported {success_count} reviews!"
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error importing file: {str(e)}")


        st.markdown(
            """
            <style>
            /* Force horizontal scroll on small screens for this specific block by preventing flex wrap on columns */
            [data-testid="stHorizontalBlock"] {
                flex-wrap: nowrap !important;
                overflow-x: auto;
                overflow-y: hidden;
                padding-bottom: 4px;
            }
            </style>
            """, unsafe_allow_html=True
        )

        hc0_v, hc0, hc0b, hc1, hc2, hc3, hc4, hc5, hc6, hc7, hc8 = st.columns([0.7, 0.6, 0.6, 1.8, 1.5, 1.2, 1.5, 1.5, 1.1, 1.5, 1.1])
        with hc0_v: st.markdown("**Sr No**")
        with hc0:  st.markdown("**ID**")
        with hc0b: st.markdown("**Rev Sr**")
        with hc1:  st.markdown("**Employee**")
        with hc2:  st.markdown("**Label / Status**")
        with hc3:  st.markdown("**Rev / Eff Date**")
        with hc4:  st.markdown("**Prev Salary**")
        with hc5:  st.markdown("**New Salary**")
        with hc6:  st.markdown("**Inc %**")
        with hc7:  st.markdown("**Increment**")
        with hc8:  st.markdown("**Action**")
        st.markdown("<hr style='margin: 0.35rem 0 0.75rem 0;'>", unsafe_allow_html=True)

        virtual_sr = 1
        for idx, row in df_filtered.iterrows():
            c0_v, c0, c0b, c1, c2, c3, c4, c5, c6, c7, c8 = st.columns([0.7, 0.6, 0.6, 1.8, 1.5, 1.2, 1.5, 1.5, 1.1, 1.5, 1.1])
            with c0_v: st.write(virtual_sr)
            with c0:  st.write(row['emp_id'])
            with c0b: st.write(str(int(row['sr_no'])) if pd.notna(row['sr_no']) else '-')
            with c1:  st.write(row['name'])
            with c2:
                label_text = row['review_name'] if row['review_name'] else '-'
                status_text = row['status'] if row['status'] else 'Proposed'
                st.markdown(f"<div style='font-size:0.9em;'>{label_text}<br><span style='color:#64748b;'>{status_text}</span></div>", unsafe_allow_html=True)
            with c3:
                r_date = format_display_date(row['review_date'])
                e_date = format_display_date(row['effective_date'])
                st.markdown(f"<div style='font-size:0.9em;'>{r_date}<br><span style='color:#64748b;'>{e_date}</span></div>", unsafe_allow_html=True)
            with c4:  st.markdown(format_currency_html(row['previous_salary'], block=True), unsafe_allow_html=True)
            with c5:  st.markdown(format_currency_html(row['new_salary'], block=True), unsafe_allow_html=True)
            with c6:  st.write(format_percentage(row['increment_percentage']))
            with c7:  st.markdown(format_currency_html(row['increment_amount'], block=True), unsafe_allow_html=True)
            with c8:
                ac1, ac2 = st.columns(2)
                with ac1:
                    if st.button("", icon=":material/edit:", key=f"s4_edit_{row['review_id']}_{idx}", type="tertiary"):
                        st.session_state.history_edit_id = row['review_id']
                        st.rerun()
                with ac2:
                    if st.button("", icon=":material/delete:", key=f"s4_del_{row['review_id']}_{idx}", type="tertiary"):
                        confirm_delete_dialog(row['review_id'])
            virtual_sr += 1

        
    else:
        st.info("No salary review logs found in the database.")


# --- EMPLOYEE PROFILES PAGE ---
elif menu == "👤 Employee Profiles":
    render_header("Employee Profiles & Salary History", "S-3")
    
    # Fetch list of all employees
    all_employees = execute_query("SELECT id, name FROM employees ORDER BY name ASC")
    
    if all_employees:
        emp_options = {emp['name']: emp['id'] for emp in all_employees}
        selected_emp_name = st.selectbox("Select Employee Profile:", list(emp_options.keys()))
        selected_emp_id = emp_options[selected_emp_name]
        
        # Load employee info
        emp_info = execute_query("SELECT * FROM employees WHERE id = ?", (selected_emp_id,))[0]
        
        # Display Department and Role
        st.markdown(f"<div style='font-size: 15px; color: #475569; margin-bottom: 15px;'><b>Department:</b> {emp_info['department'] or 'Tech'} | <b>Role:</b> {emp_info['role'] or 'N/A'}</div>", unsafe_allow_html=True)

        # Display details in cards
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(
                f"<div class='kpi-card'>"
                f"<div class='kpi-title'>Employment Status</div>"
                f"<div class='kpi-value' style='color:{'#10b981' if emp_info['status'] == 'Active' else '#f43f5e'}'>{emp_info['status']}</div>"
                f"<div class='kpi-sub-neutral'>Tenure: {format_tenure(emp_info.get('tenure'), emp_info['date_of_joining'])}</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            
        with col2:
            st.markdown(
                f"<div class='kpi-card'>"
                f"<div class='kpi-title'>Date of Joining (DoJ)</div>"
                f"<div class='kpi-value' style='font-size:22px;'>{format_display_date(emp_info['date_of_joining'])}</div>"
                f"<div class='kpi-sub-neutral'>Joining Date</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            
        with col3:
            st.markdown(
                f"<div class='kpi-card'>"
                f"<div class='kpi-title'>Joining Salary</div>"
                f"<div class='kpi-value'>{format_currency_html(emp_info['joining_salary'])}</div>"
                f"<div class='kpi-sub-neutral'>Base Starting Package</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            
        with col4:
            st.markdown(
                f"<div class='kpi-card'>"
                f"<div class='kpi-title'>Current Salary</div>"
                f"<div class='kpi-value' style='color:#4f46e5;'>{format_currency_html(emp_info['current_salary'])}</div>"
                
                f"<div class='kpi-sub'>Total Growth: "
                f"{(((emp_info['current_salary'] or 0) - (emp_info['joining_salary'] or 0)) / (emp_info['joining_salary'] or 1) * 100):+.1f}%"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            
        # Left and Right sections for Timeline Chart and History Table
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown(f"<div class='section-title'>{selected_emp_name}'s Salary Growth Curve</div>", unsafe_allow_html=True)
        
        # Load review history
        reviews = execute_query("""
            SELECT id, review_name, review_date, previous_salary, increment_amount, increment_percentage, new_salary, effective_date, remark, status 
            FROM salary_reviews 
            WHERE employee_id = ? 
            ORDER BY 
                CASE WHEN status = 'Proposed' THEN 2 ELSE 1 END,
                effective_date ASC, 
                id ASC
        """, (selected_emp_id,))
        
        if reviews:
            # Build list of points for the chart: Starting Point + Reviews
            chart_points = []
            
            # Add joining salary point
            if emp_info['joining_salary'] is not None and emp_info['date_of_joining'] is not None:
                chart_points.append({
                    'Date': emp_info['date_of_joining'],
                    'Salary': emp_info['joining_salary'],
                    'Event': 'Joining Salary',
                    'Status': 'Finalized'
                })
                
            for rev in reviews:
                # Use effective date or review date
                rev_date = rev['effective_date'] or rev['review_date'] or emp_info['date_of_joining']
                
                # Make sure we don't have None values for plot
                if rev['new_salary'] is not None:
                    chart_points.append({
                        'Date': rev_date,
                        'Salary': rev['new_salary'],
                        'Event': rev['review_name'],
                        'Status': rev['status']
                    })
            
            if chart_points:
                df_chart = pd.DataFrame(chart_points).sort_values(by='Date')
                
                # Visual Plotly Timeline
                fig = go.Figure()
                
                # Draw bars instead of lines
                fig.add_trace(go.Bar(
                    x=df_chart['Date'],
                    y=df_chart['Salary'],
                    name='Salary',
                    marker_color='#4f46e5',
                    width=0.1,
                    hovertemplate='<b>%{text}</b><br>Date: %{x}<br>Salary: ₹%{y:,.0f}<extra></extra>',
                    text=df_chart['Event'],
                    textposition='none'
                ))
                
                # Highlight proposed increments with a star above the bar
                proposed_points = df_chart[df_chart['Status'] == 'Proposed']
                if not proposed_points.empty:
                    fig.add_trace(go.Scatter(
                        x=proposed_points['Date'],
                        y=proposed_points['Salary'],
                        mode='markers',
                        name='Proposed Review',
                        marker=dict(size=14, color='#10b981', symbol='star'),
                        hovertemplate='<b>Proposed: %{text}</b><br>Date: %{x}<br>Salary: ₹%{y:,.0f}<extra></extra>',
                        text=proposed_points['Event']
                    ))
                    
                fig.update_layout(
                    plot_bgcolor='#ffffff',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridcolor='#f1f5f9', title='Timeline', type='category'),
                    yaxis=dict(showgrid=True, gridcolor='#f1f5f9', title='Salary (₹)', tickformat=',.0f'),
                    margin=dict(l=40, r=40, t=10, b=40),
                    height=350,
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig, use_container_width=True)
                
            # Chronological Table
            st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
            st.markdown("<div style='font-size:16px; font-weight:600; color:#334155; margin-bottom:10px;'>Detailed Review History</div>", unsafe_allow_html=True)
            
            df_revs = pd.DataFrame(reviews)
            
            # Format Columns
            df_revs_display = pd.DataFrame()
            df_revs_display['Review Name'] = df_revs['review_name']
            df_revs_display['Review Date'] = df_revs['review_date'].apply(format_display_date)
            df_revs_display['Effective Date'] = df_revs['effective_date'].apply(format_display_date)
            df_revs_display['Prev Salary'] = df_revs['previous_salary'].apply(lambda x: format_currency(x))
            df_revs_display['Increment'] = df_revs['increment_amount'].apply(lambda x: format_currency(x))
            df_revs_display['Increment %'] = df_revs['increment_percentage'].apply(format_percentage)
            df_revs_display['New Salary'] = df_revs['new_salary'].apply(lambda x: format_currency(x))
            df_revs_display['Status'] = df_revs['status'].apply(lambda x: f"🟢 {x}" if x == 'Finalized' else f"🟡 {x}")
            df_revs_display['Remark'] = df_revs['remark'].fillna('-')
            
            st.dataframe(df_revs_display, use_container_width=True, hide_index=True)
            
            # Export DataFrame with Date of Joining, Role, and EUR columns next to INR columns
            eur_rate = get_eur_rate()
            df_export = pd.DataFrame()
            df_export['Employee Name'] = [selected_emp_name] * len(df_revs)
            df_export['Role'] = [emp_info['role'] or 'N/A'] * len(df_revs)
            df_export['Date of Joining'] = [format_display_date(emp_info['date_of_joining'])] * len(df_revs)
            df_export['Review Name'] = df_revs['review_name']
            df_export['Review Date'] = df_revs['review_date'].apply(format_display_date)
            df_export['Effective Date'] = df_revs['effective_date'].apply(format_display_date)
            df_export['Previous Salary (INR)'] = df_revs['previous_salary'].apply(format_currency_inr)
            df_export['Previous Salary (EUR)'] = df_revs['previous_salary'].apply(lambda x: format_eur(x, eur_rate))
            df_export['Increment (INR)'] = df_revs['increment_amount'].apply(format_currency_inr)
            df_export['Increment (EUR)'] = df_revs['increment_amount'].apply(lambda x: format_eur(x, eur_rate))
            df_export['Increment %'] = df_revs['increment_percentage'].apply(format_percentage)
            df_export['New Salary (INR)'] = df_revs['new_salary'].apply(format_currency_inr)
            df_export['New Salary (EUR)'] = df_revs['new_salary'].apply(lambda x: format_eur(x, eur_rate))
            df_export['Status'] = df_revs['status']
            df_export['Remark'] = df_revs['remark'].fillna('-')
            
            excel_data = to_excel(df_export)
            st.download_button(
                label=f"📥 Export {selected_emp_name}'s History to Excel",
                data=excel_data,
                file_name=f"{selected_emp_name.replace(' ', '_')}_salary_history_{date.today().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No salary reviews recorded for this employee yet.")
            
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.info("No employees found.")


# --- RECORD NEW REVIEW PAGE ---
elif menu == "📈 Record New Review":
    
    if 'selected_emp_for_review' not in st.session_state:
        st.session_state.selected_emp_for_review = None
        
    if st.session_state.selected_emp_for_review is None:
        render_header("Record New Salary Review", "S-5")
        all_employees_raw = execute_query("""
            SELECT id, name, current_salary, department, last_review_date, last_review_effective_date 
            FROM employees 
            WHERE status = 'Active' 
            AND id NOT IN (SELECT employee_id FROM salary_reviews WHERE status = 'Proposed')
            ORDER BY name ASC
        """)
        
        if all_employees_raw:
            st.markdown("<div class='content-section'>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Select Employee for Review</div>", unsafe_allow_html=True)
            
            # Determine available years for the filter
            available_years = set()
            for emp in all_employees_raw:
                if emp['last_review_effective_date']:
                    available_years.add(str(emp['last_review_effective_date'])[:4])
                    
            years_list = sorted(list(available_years), reverse=True)
            filter_options = ["All", "> 12 Months Ago", "Current Year", "Previous Years"] + years_list + ["Blank (No Review)"]

            # Initialise session state for filter on first load only
            if "s5_year_filter" not in st.session_state:
                st.session_state.s5_year_filter = ["All"]

            col_f1, col_f2 = st.columns(2)
            with col_f1:
                selected_year_filters = st.multiselect(
                    "Year of Last Review Effective Date (Recent Review Submissions are EXCLUDED)",
                    options=filter_options,
                    default=st.session_state.s5_year_filter,
                    key="s5_year_filter",
                    placeholder="Select one or more..."
                )
            # Default to All if nothing selected
            if not selected_year_filters:
                selected_year_filters = ["All"]

            current_year_str = str(date.today().year)
            _today = date.today()
            _m = _today.month - 11
            _y = _today.year + (_m - 1) // 12
            _m = (_m - 1) % 12 + 1
            twelve_months_ago = _today.replace(year=_y, month=_m)
            all_employees = []
            for emp in all_employees_raw:
                matched = False
                for selected_year_filter in selected_year_filters:
                    if matched:
                        break
                    if selected_year_filter == "All":
                        matched = True
                    elif selected_year_filter == "> 12 Months Ago":
                        if not emp['last_review_effective_date']:
                            matched = True  # No review = definitely overdue
                        else:
                            try:
                                eff_date = datetime.strptime(emp['last_review_effective_date'], "%Y-%m-%d").date()
                                if eff_date < twelve_months_ago:
                                    matched = True
                            except:
                                pass
                    elif selected_year_filter == "Blank (No Review)":
                        if not emp['last_review_effective_date']:
                            matched = True
                    elif selected_year_filter == "Current Year":
                        if emp['last_review_effective_date'] and emp['last_review_effective_date'].startswith(current_year_str):
                            matched = True
                    elif selected_year_filter == "Previous Years":
                        if emp['last_review_effective_date']:
                            emp_year = emp['last_review_effective_date'][:4]
                            if emp_year.isdigit() and int(emp_year) < int(current_year_str):
                                matched = True
                    else:
                        if emp['last_review_effective_date'] and emp['last_review_effective_date'].startswith(selected_year_filter):
                            matched = True
                if matched:
                    all_employees.append(emp)
            
            # Table Header
            hcol0, hcol1, hcol2, hcol3, hcol4, hcol5, hcol6 = st.columns([0.8, 2.5, 1.5, 1.5, 1.5, 1.5, 1.5])
            with hcol0: st.markdown("**Emp ID**")
            with hcol1: st.markdown("**Employee Name**")
            with hcol2: st.markdown("**Department**")
            with hcol3: st.markdown("**Current<br>Salary**", unsafe_allow_html=True)
            with hcol4: st.markdown("**Last Rev Date**")
            with hcol5: st.markdown("**Effective Date**")
            with hcol6: st.markdown("**Action**")
            st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
            
            for emp in all_employees:
                col0, col1, col2, col3, col4, col5, col6 = st.columns([0.8, 2.5, 1.5, 1.5, 1.5, 1.5, 1.5])
                with col0: st.write(emp['id'])
                with col1: st.write(emp['name'])
                with col2: st.write(emp['department'] or "Tech")
                with col3: st.markdown(format_currency_html(emp['current_salary'], block=True), unsafe_allow_html=True)
                with col4: st.write(format_display_date(emp['last_review_date']))
                with col5: st.write(format_display_date(emp['last_review_effective_date']))
                with col6:
                    if st.button("📝", key=f"rec_btn_{emp['id']}", help="Record Review"):
                        st.session_state.selected_emp_for_review = emp['id']
                        st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        else:
            st.info("No active employees found to record reviews for.")
            
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        recent_reviews = execute_query("""
            SELECT e.id as emp_id, e.name, e.date_of_joining, e.role, r.review_name, r.review_date, r.previous_salary, r.increment_amount, r.new_salary, r.status
            FROM salary_reviews r
            JOIN employees e ON r.employee_id = e.id
            ORDER BY r.id DESC LIMIT 10
        """)
        col_sec_title, col_sec_exp = st.columns([3, 1])
        with col_sec_title:
            st.markdown("<div class='section-title'>Recent Review Submissions</div>", unsafe_allow_html=True)
        with col_sec_exp:
            if recent_reviews:
                df_rec = pd.DataFrame(recent_reviews)
                eur_rate = get_eur_rate()
                df_rec_display = pd.DataFrame()
                df_rec_display['Emp ID'] = df_rec['emp_id']
                df_rec_display['Employee Name'] = df_rec['name']
                df_rec_display['Role'] = df_rec['role'].fillna('-')
                df_rec_display['Date of Joining'] = df_rec['date_of_joining'].apply(format_display_date)
                df_rec_display['Review Name'] = df_rec['review_name']
                df_rec_display['Date'] = df_rec['review_date'].apply(format_display_date)
                
                df_rec_display['Previous Salary (INR)'] = df_rec['previous_salary'].apply(format_currency_inr)
                df_rec_display['Previous Salary (EUR)'] = df_rec['previous_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_rec_display['Increment (INR)'] = df_rec['increment_amount'].apply(format_currency_inr)
                df_rec_display['Increment (EUR)'] = df_rec['increment_amount'].apply(lambda x: format_eur(x, eur_rate))
                df_rec_display['New Salary (INR)'] = df_rec['new_salary'].apply(format_currency_inr)
                df_rec_display['New Salary (EUR)'] = df_rec['new_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_rec_display['Status'] = df_rec['status']
                
                st.download_button(
                    label="📥 Export to XLSX",
                    data=to_excel(df_rec_display),
                    file_name=f"recent_reviews_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        if recent_reviews:
            st.dataframe(df_rec_display, use_container_width=True, hide_index=True)
        else:
            st.info("No salary reviews recorded yet.")
        st.markdown("</div>", unsafe_allow_html=True)
        
    else:
        render_header("Record New Salary Review", "S-9")
        components.html("<script>window.parent.scrollTo(0, 0);</script>", height=0)
        
        # Form for selected employee
        emp_id = st.session_state.selected_emp_for_review
        emp_info = execute_query("SELECT id, name, current_salary FROM employees WHERE id = ?", (emp_id,))[0]
        
        col_btn1, col_btn2 = st.columns([1, 4])
        with col_btn1:
            if st.button("← Back to Employee List"):
                st.session_state.selected_emp_for_review = None
                st.rerun()
        with col_btn2:
            if st.button("🔄 Refresh Data"):
                st.rerun()
            
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown(f"<div class='section-title'>Review Details for {emp_info['name']}</div>", unsafe_allow_html=True)
        
        cur_sal = emp_info['current_salary']
        st.info(f"Current salary for this employee: **{format_currency(cur_sal)}**")
        
        col1, col2 = st.columns(2)
        with col1:
            review_name = st.text_input("Review Label / Name:", value="Review-26", placeholder="e.g. Review-26, Aug-25 Review", key=f"rev_name_{emp_id}")
            review_date = st.date_input("Review Date (Date of decision):", value=date.today(), key=f"rev_date_{emp_id}", format="DD-MM-YYYY")
            effective_date = st.date_input("Effective Date (Date salary takes effect):", value=date.today(), key=f"eff_date_{emp_id}", format="DD-MM-YYYY")
        
        with col2:
            inc_type = st.radio("Increment Input Type:", ["Absolute Value", "Percentage Increase"], key=f"inc_type_{emp_id}")
            if inc_type == "Absolute Value":
                inc_amt = st.number_input("Increment Amount (₹):", min_value=0.0, step=5000.0, value=50000.0, key=f"inc_amt_{emp_id}")
                if inc_amt > 0:
                    st.caption(f"💶 **EUR:** {format_eur(inc_amt, get_eur_rate())} | ✍️ {num_to_words_indian(inc_amt)}")
                inc_pct = (inc_amt / cur_sal * 100.0) if cur_sal > 0 else 0.0
            else:
                inc_pct = st.number_input("Increment Percentage (%):", min_value=0.0, max_value=200.0, step=0.5, value=5.0, key=f"inc_pct_{emp_id}")
                inc_amt = (inc_pct / 100.0) * cur_sal
                if inc_amt > 0:
                    st.caption(f"💶 **EUR Equivalent:** {format_eur(inc_amt, get_eur_rate())} | ✍️ {num_to_words_indian(inc_amt)}")
            
            st.write(f"Calculated Increment Amount: **{format_currency(inc_amt)}** (In Words: {num_to_words_indian(inc_amt)})")
            st.write(f"Calculated Increment Percentage: **{inc_pct:.2f}%**")
            st.write(f"Resulting Salary: **{format_currency(cur_sal + inc_amt)}** (In Words: {num_to_words_indian(cur_sal + inc_amt)})")
            
        remark = st.text_area("Remarks / Notes:", placeholder="Add review feedback, offer letter details, etc.", key=f"remark_{emp_id}")
        status_opt = st.radio("Status of this review:", ["Proposed (Draft for planner)", "Finalized (Apply to current salary immediately)"], horizontal=True, key=f"status_{emp_id}")
        
        status = "Proposed" if status_opt.startswith("Proposed") else "Finalized"
        
        submit = st.button("Save Salary Review", type="primary", key=f"submit_{emp_id}")
        
        if submit:
            new_salary = cur_sal + inc_amt
            
            existing_proposed = execute_query(
                "SELECT id FROM salary_reviews WHERE employee_id = ? AND review_name = ? AND status = 'Proposed'",
                (emp_id, review_name)
            )
            
            if existing_proposed:
                execute_query("""
                    UPDATE salary_reviews 
                    SET review_date = ?, previous_salary = ?, increment_amount = ?, increment_percentage = ?, new_salary = ?, effective_date = ?, remark = ?, status = ?
                    WHERE id = ?
                """, (
                    review_date.strftime("%Y-%m-%d"), cur_sal, inc_amt, inc_pct, new_salary, 
                    effective_date.strftime("%Y-%m-%d"), remark, status, existing_proposed[0]['id']
                ), commit=True)
            else:
                execute_query("""
                    INSERT INTO salary_reviews (employee_id, review_name, review_date, previous_salary, increment_amount, increment_percentage, new_salary, effective_date, remark, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    emp_id, review_name, review_date.strftime("%Y-%m-%d"), cur_sal, inc_amt, inc_pct, new_salary, 
                    effective_date.strftime("%Y-%m-%d"), remark, status
                ), commit=True)
            
            if status == "Finalized":
                execute_query("UPDATE employees SET current_salary = ?, last_review_date = ?, last_review_effective_date = ? WHERE id = ?", 
                              (new_salary, review_date.strftime("%Y-%m-%d"), effective_date.strftime("%Y-%m-%d"), emp_id), commit=True)
                st.success(f"Salary review finalized! {emp_info['name']}'s current salary has been updated to {format_currency(new_salary)}.")
            else:
                st.success(f"Salary review saved as 'Proposed' draft in the Review Planner.")
            
            st.session_state.selected_emp_for_review = None
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)


# --- REVIEW PLANNER PAGE (Review-26) ---
elif menu == "📅 Review Planner (Review-26)":
    if 'selected_review_to_edit' not in st.session_state:
        st.session_state.selected_review_to_edit = None
        
    if st.session_state.selected_review_to_edit is None:
        render_header("Salary Review Planner (Review-26)", "S-6")
    else:
        render_header("Salary Review Planner (Review-26)", "S-10")
    
    st.markdown(
        "<div style='background-color:#eff6ff; border: 1px solid #bfdbfe; border-radius:12px; padding:15px; margin-bottom:20px; color:#1e40af; font-size:14px;'>"
        "💡 **Review Planner Context**: This workspace lists reviews marked as **Proposed**. These drafts represent the upcoming cycle. "
        "You can inspect them, calculate budget implications, edit individual amounts, and **Bulk Approve** them to finalize the salaries in one click."
        "</div>",
        unsafe_allow_html=True
    )
    
    # Load all proposed reviews
    proposed_reviews = execute_query("""
        SELECT r.id as review_id, e.id as emp_id, e.name, e.date_of_joining, e.role, e.current_salary as live_salary, r.previous_salary, r.increment_amount, r.increment_percentage, r.new_salary, r.review_date, r.effective_date, r.remark 
        FROM salary_reviews r
        JOIN employees e ON r.employee_id = e.id
        WHERE r.status = 'Proposed'
        ORDER BY e.name ASC
    """)
    
    if proposed_reviews:
        df_prop = pd.DataFrame(proposed_reviews)
        
        # Display Stats Summary of Proposed changes
        total_p_inc = df_prop['increment_amount'].sum()
        avg_p_inc = df_prop['increment_percentage'].mean()
        
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Employees with Proposed Review", len(df_prop))
        with c2:
            st.metric("Total Increments Budget Impact", format_currency(total_p_inc))
        with c3:
            st.metric("Average Increment %", f"{avg_p_inc:.2f}%")
            
        if st.session_state.selected_review_to_edit is None:
            # Table list
            st.markdown("<div class='content-section'>", unsafe_allow_html=True)
            c_title, c_exp = st.columns([3, 1])
            with c_title:
                st.markdown("<div class='section-title'>Proposed Reviews List</div>", unsafe_allow_html=True)
            with c_exp:
                eur_rate = get_eur_rate()
                df_export = df_prop.copy()
                df_export['Role'] = df_export['role'].fillna('-')
                df_export['Date of Joining'] = df_export['date_of_joining'].apply(format_display_date)
                
                df_export['Previous Salary (INR)'] = df_export['previous_salary'].apply(format_currency_inr)
                df_export['Previous Salary (EUR)'] = df_export['previous_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_export['Increment (INR)'] = df_export['increment_amount'].apply(format_currency_inr)
                df_export['Increment (EUR)'] = df_export['increment_amount'].apply(lambda x: format_eur(x, eur_rate))
                df_export['Increment %'] = df_export['increment_percentage'].apply(format_percentage)
                df_export['New Salary (INR)'] = df_export['new_salary'].apply(format_currency_inr)
                df_export['New Salary (EUR)'] = df_export['new_salary'].apply(lambda x: format_eur(x, eur_rate))
                
                df_export = df_export[['emp_id', 'name', 'Role', 'Date of Joining', 
                                       'Previous Salary (INR)', 'Previous Salary (EUR)', 
                                       'Increment (INR)', 'Increment (EUR)', 'Increment %', 
                                       'New Salary (INR)', 'New Salary (EUR)']].rename(columns={'emp_id': 'Emp ID', 'name': 'Employee Name'})
                st.download_button(
                    label="📥 Export to XLSX",
                    data=to_excel(df_export),
                    file_name=f"proposed_reviews_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            # Custom Table with Action Buttons
            hcol0, hcol1, hcol2, hcol3, hcol4, hcol5, hcol6 = st.columns([0.8, 2, 1.5, 1.5, 1, 1.5, 1.5])
            with hcol0: st.markdown("**Emp ID**")
            with hcol1: st.markdown("**Employee Name**")
            with hcol2: st.markdown("**Current<br>Salary**", unsafe_allow_html=True)
            with hcol3: st.markdown("**Proposed Inc.**")
            with hcol4: st.markdown("**Inc. %**")
            with hcol5: st.markdown("**New<br>Salary**", unsafe_allow_html=True)
            with hcol6: st.markdown("**Action**")
            st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
            
            for idx, row in df_prop.iterrows():
                col0, col1, col2, col3, col4, col5, col6 = st.columns([0.8, 2, 1.5, 1.5, 1, 1.5, 1.5])
                with col0: st.write(row['emp_id'])
                with col1: st.write(row['name'])
                with col2: st.markdown(format_currency_html(row['previous_salary'], block=True), unsafe_allow_html=True)
                with col3: st.markdown(format_currency_html(row['increment_amount'], block=True), unsafe_allow_html=True)
                with col4: st.write(f"{row['increment_percentage']:.2f}%")
                with col5: st.markdown(format_currency_html(row['new_salary'], block=True), unsafe_allow_html=True)
                with col6:
                    if st.button("✏️", key=f"adj_btn_{row['review_id']}", help="Adjust review"):
                        st.session_state.selected_review_to_edit = row['review_id']
                        st.rerun()
                        
            st.markdown("<div style='margin-top:15px;'></div>", unsafe_allow_html=True)
            
            # Bulk Approve
            st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
            col_btn1, col_btn2 = st.columns([1, 4])
            with col_btn1:
                approve_all = st.button("🔥 Finalize All Proposed", type="primary")
                if approve_all:
                    for row_dict in proposed_reviews:
                        execute_query("UPDATE salary_reviews SET status = 'Finalized' WHERE id = ?", (row_dict['review_id'],), commit=True)
                        execute_query("UPDATE employees SET current_salary = ?, last_review_date = ?, last_review_effective_date = ? WHERE id = ?", 
                                      (row_dict['new_salary'], row_dict['review_date'], row_dict['effective_date'], row_dict['emp_id']), commit=True)
                    st.success("All proposed reviews have been successfully finalized and applied to current salaries!")
                    st.rerun()
            with col_btn2:
                st.write("")
            st.markdown("</div>", unsafe_allow_html=True)
            
        else:
            # Individual Editing form
            components.html("<script>window.parent.scrollTo(0, 0);</script>", height=0)
            review_id = st.session_state.selected_review_to_edit
            filtered_df = df_prop[df_prop['review_id'] == review_id]
            
            if filtered_df.empty:
                st.warning("Review not found. It may have been finalized or deleted.")
                if st.button("← Back to Proposed Reviews List"):
                    st.session_state.selected_review_to_edit = None
                    st.rerun()
                st.stop()
                
            row_selected = filtered_df.iloc[0]
            
            col_btn1, col_btn2 = st.columns([1, 4])
            with col_btn1:
                if st.button("← Back to Proposed Reviews List"):
                    st.session_state.selected_review_to_edit = None
                    st.rerun()
            with col_btn2:
                if st.button("🔄 Refresh Data"):
                    live_sal = execute_query("SELECT current_salary FROM employees WHERE id = ?", (int(row_selected['emp_id']),))[0]['current_salary']
                    if float(live_sal) != float(row_selected['previous_salary']):
                        new_sal = float(live_sal) + float(row_selected['increment_amount'])
                        pct = (float(row_selected['increment_amount']) / float(live_sal) * 100.0) if float(live_sal) > 0 else 0.0
                        execute_query("UPDATE salary_reviews SET previous_salary = ?, increment_percentage = ?, new_salary = ? WHERE id = ?", 
                                      (float(live_sal), pct, new_sal, int(review_id)), commit=True)
                        st.success("Refreshed! Calculations updated with the latest base salary.")
                    st.rerun()
                
            st.markdown("<div class='content-section'>", unsafe_allow_html=True)
            st.markdown(f"<div class='section-title'>Adjust Proposed Review for {row_selected['name']}</div>", unsafe_allow_html=True)
            
            st.info(f"Current salary: **{format_currency(row_selected['previous_salary'])}**")
            
            with st.form("edit_proposed_form"):
                col_edit1, col_edit2 = st.columns(2)
                with col_edit1:
                    edit_inc = st.number_input("Adjust Increment Amount (₹):", min_value=0.0, value=float(row_selected['increment_amount']), step=5000.0)
                    if edit_inc > 0:
                        st.caption(f"💶 **EUR:** {format_eur(edit_inc, get_eur_rate())} | ✍️ {num_to_words_indian(edit_inc)}")
                    edit_eff = st.date_input("Adjust Effective Date:", value=datetime.strptime(row_selected['effective_date'], "%Y-%m-%d").date() if row_selected['effective_date'] else date.today(), format="DD-MM-YYYY")
                with col_edit2:
                    edit_rem = st.text_area("Adjust Remark:", value=str(row_selected['remark'] or ""))
                
                save_adj = st.form_submit_button("Save Adjustments")
                
                if save_adj:
                    new_sal = row_selected['previous_salary'] + edit_inc
                    pct = (edit_inc / row_selected['previous_salary'] * 100.0) if row_selected['previous_salary'] > 0 else 0.0
                    
                    execute_query("""
                        UPDATE salary_reviews 
                        SET increment_amount = ?, increment_percentage = ?, new_salary = ?, effective_date = ?, remark = ?
                        WHERE id = ?
                    """, (edit_inc, pct, new_sal, edit_eff.strftime("%Y-%m-%d"), edit_rem, review_id), commit=True)
                    st.success(f"Adjustments saved for {row_selected['name']}!")
                    st.session_state.selected_review_to_edit = None
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        
    else:
        st.info("No proposed reviews currently in draft status. You can add a proposed review via the 'Record New Review' page.")


# --- MANAGE EMPLOYEES PAGE ---
elif menu == "⚙️ Manage Employees":

    if 'manage_emp_action' not in st.session_state:
        st.session_state.manage_emp_action = 'list'

    def clean_employee_text(value, fallback="N/A"):
        if value is None or pd.isna(value):
            return fallback
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return fallback
        return text
        
    if st.session_state.manage_emp_action == 'add':
        render_header("Register New Employee", "S-12")
        components.html("<script>window.parent.scrollTo(0, 0);</script>", height=0)
        
        if st.button("← Back to Employee Table"):
            st.session_state.manage_emp_action = 'list'
            st.rerun()
            
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>Register New Employee</div>", unsafe_allow_html=True)
        with st.form("add_employee_form"):
            new_name = st.text_input("Full Name:", placeholder="First Last")
            new_dept = st.text_input("Department:", value="Tech")
            new_role = st.text_input("Role:", placeholder="e.g. Developer, Designer")
            new_start = st.date_input("Date of Joining (DoJ):", value=date.today(), format="DD-MM-YYYY")
            new_sal = st.number_input("Joining Base Salary (₹):", min_value=0.0, step=10000.0, value=500000.0)
            if new_sal > 0:
                st.caption(f"💶 **EUR Equivalent:** {format_eur(new_sal, get_eur_rate())} | ✍️ **In Words:** {num_to_words_indian(new_sal)}")
            
            add_submit = st.form_submit_button("Register Employee", type="primary")
            
            if add_submit:
                if not new_name.strip():
                    st.error("Employee Name is required!")
                else:
                    res = execute_query("""
                        INSERT OR IGNORE INTO employees (name, department, date_of_joining, role, joining_salary, current_salary, status)
                        VALUES (?, ?, ?, ?, ?, ?, 'Active')
                    """, (new_name.strip(), new_dept.strip(), new_start.strftime("%Y-%m-%d"), new_role.strip() if new_role.strip() else None, new_sal, new_sal), commit=True)
                    
                    if res:
                        st.success(f"Registered {new_name} successfully!")
                        st.session_state.manage_emp_action = 'list'
                        st.rerun()
                    else:
                        st.error("Employee name already exists in the database.")
        st.markdown("</div>", unsafe_allow_html=True)

    elif st.session_state.manage_emp_action != 'list':
        render_header("Edit Employee Details", "S-11")
        edit_id = st.session_state.manage_emp_action
        emp_rows = execute_query("""
            SELECT id, name, department, role, date_of_joining, joining_salary, current_salary, status, last_review_date, last_review_effective_date, resign_date, lwd
            FROM employees
            WHERE id = ?
        """, (edit_id,))

        if not emp_rows:
            st.warning("Employee record was not found.")
            if st.button("← Back to Employee Table"):
                st.session_state.manage_emp_action = 'list'
                st.rerun()
        else:
            emp_data = emp_rows[0]

            if st.button("← Back to Employee Table"):
                st.session_state.manage_emp_action = 'list'
                st.rerun()

            st.markdown("<div class='content-section'>", unsafe_allow_html=True)
            st.markdown(f"<div class='section-title'>Edit Employee: {emp_data['name']}</div>", unsafe_allow_html=True)

            rev_key = f"edit_rev_{edit_id}"
            eff_key = f"edit_eff_{edit_id}"
            res_key = f"edit_res_{edit_id}"
            lwd_key = f"edit_lwd_{edit_id}"
            if rev_key not in st.session_state:
                st.session_state[rev_key] = datetime.strptime(emp_data['last_review_date'], "%Y-%m-%d").date() if emp_data['last_review_date'] else None
            if eff_key not in st.session_state:
                st.session_state[eff_key] = datetime.strptime(emp_data['last_review_effective_date'], "%Y-%m-%d").date() if emp_data['last_review_effective_date'] else None
            if res_key not in st.session_state:
                st.session_state[res_key] = datetime.strptime(emp_data['resign_date'], "%Y-%m-%d").date() if emp_data.get('resign_date') else None
            if lwd_key not in st.session_state:
                st.session_state[lwd_key] = datetime.strptime(emp_data['lwd'], "%Y-%m-%d").date() if emp_data.get('lwd') else None

            with st.form("manage_employee_full_page_edit"):
                r1c1, r1c2, r1c_role, r1c3 = st.columns(4)
                with r1c1:
                    edit_name = st.text_input("Full Name:", value=emp_data['name'] or "")
                with r1c2:
                    edit_dept = st.text_input("Department:", value=clean_employee_text(emp_data['department'], ""))
                with r1c_role:
                    edit_role = st.text_input("Role:", value=clean_employee_text(emp_data.get('role'), ""))
                with r1c3:
                    edit_status = st.selectbox("Status:", ["Active", "Resigned"],
                        index=0 if emp_data['status'] == 'Active' else 1)

                r2c1, r2c2, r2c3 = st.columns(3)
                with r2c1:
                    edit_start = st.date_input("Date of Joining (DoJ):",
                        value=datetime.strptime(emp_data['date_of_joining'], "%Y-%m-%d").date() if emp_data['date_of_joining'] else date.today(),
                        format="DD-MM-YYYY")
                with r2c2:
                    edit_joining_salary = st.number_input("Joining Salary (₹):", min_value=0.0, step=10000.0,
                        value=float(emp_data['joining_salary'] or 0))
                    if edit_joining_salary > 0:
                        st.caption(f"💶 **EUR:** {format_eur(edit_joining_salary, get_eur_rate())} | ✍️ {num_to_words_indian(edit_joining_salary)}")
                with r2c3:
                    edit_current_salary = st.number_input("Current Salary (₹):", min_value=0.0, step=10000.0,
                        value=float(emp_data['current_salary'] or 0))
                    if edit_current_salary > 0:
                        st.caption(f"💶 **EUR:** {format_eur(edit_current_salary, get_eur_rate())} | ✍️ {num_to_words_indian(edit_current_salary)}")

                r3c1, r3c2, r3c3 = st.columns(3)
                with r3c1:
                    edit_last_rev = st.date_input("Last Review Date:",
                        value=st.session_state[rev_key], key=rev_key, format="DD-MM-YYYY")
                with r3c2:
                    edit_last_eff = st.date_input("Last Review Effective Date:",
                        value=st.session_state[eff_key], key=eff_key, format="DD-MM-YYYY")
                with r3c3:
                    edit_resign_date = st.date_input("Resign Date:",
                        value=st.session_state[res_key], key=res_key, format="DD-MM-YYYY")

                r4c1, r4c2, r4c3 = st.columns(3)
                with r4c1:
                    edit_lwd = st.date_input("Last Working Day (LWD):",
                        value=st.session_state[lwd_key], key=lwd_key, format="DD-MM-YYYY")

                edit_submit = st.form_submit_button("💾 Save Employee", type="primary")

                if edit_submit:
                    if not edit_name.strip():
                        st.error("Employee name is required.")
                    else:
                        execute_query("""
                            UPDATE employees
                            SET name = ?, department = ?, role = ?, date_of_joining = ?, joining_salary = ?, current_salary = ?, status = ?, last_review_date = ?, last_review_effective_date = ?, resign_date = ?, lwd = ?
                            WHERE id = ?
                        """, (
                            edit_name.strip(),
                            edit_dept.strip(),
                            edit_role.strip() if edit_role.strip() else None,
                            edit_start.strftime("%Y-%m-%d"),
                            edit_joining_salary,
                            edit_current_salary,
                            edit_status,
                            st.session_state.get(rev_key).strftime("%Y-%m-%d") if st.session_state.get(rev_key) else None,
                            st.session_state.get(eff_key).strftime("%Y-%m-%d") if st.session_state.get(eff_key) else None,
                            st.session_state.get(res_key).strftime("%Y-%m-%d") if st.session_state.get(res_key) else None,
                            st.session_state.get(lwd_key).strftime("%Y-%m-%d") if st.session_state.get(lwd_key) else None,
                            edit_id
                        ), commit=True)
                        st.success(f"Updated {edit_name.strip()}'s employee profile.")
                        st.session_state.manage_emp_action = 'list'
                        st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)


    else:
        render_header("Manage Employees", "S-7")
        
        col_title, col_add = st.columns([3, 1])
        with col_add:
            if st.button("➕ Add New Employee", type="primary", use_container_width=True):
                st.session_state.manage_emp_action = 'add'
                st.rerun()

        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>Employee Table</div>", unsafe_allow_html=True)
        
        all_emps = execute_query("""
            SELECT id, name, department, role, date_of_joining, joining_salary, current_salary, status, last_review_date, last_review_effective_date, resign_date, lwd
            FROM employees
            ORDER BY name ASC
        """)
        if all_emps:
            df_manage = pd.DataFrame(all_emps)
            search_col, status_col = st.columns([2, 1])
            with search_col:
                search_text = st.text_input("Search employee:", placeholder="Type a name, department or role")
            with status_col:
                status_filter = st.selectbox("Status:", ["All", "Active", "Resigned"])
            
            if search_text.strip():
                needle = search_text.strip().lower()
                df_manage = df_manage[
                    df_manage['name'].fillna("").str.lower().str.contains(needle)
                    | df_manage['department'].fillna("").str.lower().str.contains(needle)
                    | df_manage['role'].fillna("").str.lower().str.contains(needle)
                ]
            if status_filter != "All":
                df_manage = df_manage[df_manage['status'] == status_filter]
            
            col_stat1, col_stat2 = st.columns([3, 1])
            with col_stat1:
                st.write(f"Showing **{len(df_manage)}** employees.")
            with col_stat2:
                eur_rate = get_eur_rate()
                df_export = df_manage.copy()
                df_export['date_of_joining'] = df_export['date_of_joining'].apply(format_display_date)
                df_export['joining_salary_inr'] = df_export['joining_salary'].apply(format_currency_inr)
                df_export['joining_salary_eur'] = df_export['joining_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_export['current_salary_inr'] = df_export['current_salary'].apply(format_currency_inr)
                df_export['current_salary_eur'] = df_export['current_salary'].apply(lambda x: format_eur(x, eur_rate))
                df_export['last_review_date'] = df_export['last_review_date'].apply(format_display_date)
                df_export['last_review_effective_date'] = df_export['last_review_effective_date'].apply(format_display_date)
                df_export['resign_date'] = df_export['resign_date'].apply(format_display_date)
                df_export['lwd'] = df_export['lwd'].apply(format_display_date)
                df_export = df_export.rename(columns={
                    'id': 'Emp ID', 'name': 'Employee Name', 'department': 'Department',
                    'role': 'Role', 'date_of_joining': 'Date of Joining',
                    'joining_salary_inr': 'Joining Salary (INR)', 'joining_salary_eur': 'Joining Salary (EUR)',
                    'current_salary_inr': 'Current Salary (INR)', 'current_salary_eur': 'Current Salary (EUR)',
                    'status': 'Status', 'last_review_date': 'Last Review Date',
                    'last_review_effective_date': 'Last Review Effective Date',
                    'resign_date': 'Resigned Date', 'lwd': 'Last Working Day'
                })
                df_export = df_export[['Emp ID', 'Employee Name', 'Department', 'Role', 'Date of Joining', 
                                       'Joining Salary (INR)', 'Joining Salary (EUR)', 'Current Salary (INR)', 'Current Salary (EUR)', 
                                       'Status', 'Last Review Date', 'Last Review Effective Date', 'Resigned Date', 'Last Working Day']]
                st.download_button(
                    label="📥 Export to XLSX",
                    data=to_excel(df_export),
                    file_name=f"manage_employees_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            hcol0, hcol1, hcol2, hcol3, hcol4, hcol5, hcol6, hcol7, hcol8, hcol9, hcol10, hcol11 = st.columns([0.4, 0.4, 1.5, 1.0, 1.0, 1.0, 1.0, 1.2, 1.0, 1.0, 0.8, 0.8])
            with hcol0: st.markdown("**Sr No**")
            with hcol1: st.markdown("**ID**")
            with hcol2: st.markdown("**Name**")
            with hcol3: st.markdown("**Dept**")
            with hcol4: st.markdown("**DoJ**")
            with hcol5: st.markdown("**Join Sal**")
            with hcol6: st.markdown("**Cur Sal**")
            with hcol7: st.markdown("**Last Rev/Eff**")
            with hcol8: st.markdown("**Resign Dt**")
            with hcol9: st.markdown("**LWD**")
            with hcol10: st.markdown("**Status**")
            with hcol11: st.markdown("**Action**")
            st.markdown("<hr style='margin: 0.35rem 0 0.75rem 0;'>", unsafe_allow_html=True)
            
            virtual_sr = 1
            for idx, row in df_manage.iterrows():
                col0, col1, col2, col3, col4, col5, col6, col7, col8, col9, col10, col11 = st.columns([0.4, 0.4, 1.5, 1.0, 1.0, 1.0, 1.0, 1.2, 1.0, 1.0, 0.8, 0.8])
                with col0: st.write(virtual_sr)
                with col1: st.write(row['id'])
                with col2:
                    name_val = row['name']
                    role_val = row.get('role') or 'N/A'
                    st.markdown(f"<div style='font-weight:600;'>{name_val}</div><div style='font-size:0.85em; color:#64748b;'>{role_val}</div>", unsafe_allow_html=True)
                with col3: st.write(clean_employee_text(row['department']))
                with col4: st.write(format_display_date(row['date_of_joining']))
                with col5: st.markdown(format_currency_html(row['joining_salary'], block=True), unsafe_allow_html=True)
                with col6: st.markdown(format_currency_html(row['current_salary'], block=True), unsafe_allow_html=True)
                with col7:
                    r_date = format_display_date(row['last_review_date'])
                    e_date = format_display_date(row['last_review_effective_date'])
                    st.markdown(f"<div style='font-size:0.9em;'>{r_date}<br><span style='color:#64748b;'>{e_date}</span></div>", unsafe_allow_html=True)
                with col8: st.write(format_display_date(row.get('resign_date')))
                with col9: st.write(format_display_date(row.get('lwd')))
                with col10:
                    st.write(row['status'])
                with col11:
                    if st.button("✏️", key=f"manage_emp_edit_{row['id']}_{idx}", help="Edit employee"):
                        st.session_state.manage_emp_action = int(row['id'])
                        st.rerun()
                st.markdown("<hr style='margin: 0.2rem 0 0.45rem 0; opacity: 0.35;'>", unsafe_allow_html=True)
                virtual_sr += 1
        else:
            st.info("No employees in database.")
        st.markdown("</div>", unsafe_allow_html=True)
            
        # Master Backup Export Section
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>System Database Backup</div>", unsafe_allow_html=True)
        st.write("Export the entire system database (all employees and all reviews) as a multi-sheet Excel file for backup or audit purposes.")
        
        db_excel = db_to_excel()
        st.download_button(
            label="📥 Export Master Database (All Tables) to Excel",
            data=db_excel,
            file_name=f"salary_review_master_backup_{date.today().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.markdown("</div>", unsafe_allow_html=True)

# --- BONUSES PAGE ---
elif menu == "🎁 Bonuses":
    render_header("Manage Employee Bonuses", "S-13")
    
    if 'bonus_action' not in st.session_state:
        st.session_state.bonus_action = 'list'
        
    if st.session_state.bonus_action == 'add' or isinstance(st.session_state.bonus_action, int):
        is_edit = isinstance(st.session_state.bonus_action, int)
        action_title = "Edit Bonus" if is_edit else "Add New Bonus"
        
        components.html("<script>window.parent.scrollTo(0, 0);</script>", height=0)
        
        if st.button("← Back to Bonuses List"):
            st.session_state.bonus_action = 'list'
            st.rerun()
            
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        st.markdown(f"<div class='section-title'>{action_title}</div>", unsafe_allow_html=True)
        
        # Load employees for dropdown
        all_emps = execute_query("SELECT id, name FROM employees ORDER BY name ASC")
        emp_options = {emp['name']: emp['id'] for emp in all_emps} if all_emps else {}
        
        # Defaults
        def_emp_id = list(emp_options.values())[0] if emp_options else None
        def_amount = 50000.0
        def_date_dec = date.today()
        def_due_date = date.today()
        
        if is_edit:
            bonus_id = st.session_state.bonus_action
            existing = execute_query("SELECT * FROM bonuses WHERE id = ?", (bonus_id,))
            if existing:
                row = existing[0]
                def_emp_id = row['employee_id']
                def_amount = float(row['amount'])
                def_date_dec = datetime.strptime(row['date_declared'], "%Y-%m-%d").date() if row['date_declared'] else date.today()
                def_due_date = datetime.strptime(row['due_date'], "%Y-%m-%d").date() if row['due_date'] else date.today()
                
        # Find default emp index
        emp_idx = 0
        if def_emp_id in emp_options.values():
            emp_idx = list(emp_options.values()).index(def_emp_id)
            
        with st.form("bonus_form"):
            col1, col2 = st.columns(2)
            with col1:
                sel_emp_name = st.selectbox("Employee:", list(emp_options.keys()), index=emp_idx)
                amt = st.number_input("Bonus Amount (₹):", min_value=0.0, value=def_amount, step=10000.0)
                if amt > 0:
                    st.caption(f"💶 **EUR Equivalent:** {format_eur(amt, get_eur_rate())} | ✍️ **In Words:** {num_to_words_indian(amt)}")
            with col2:
                date_dec = st.date_input("Date Declared:", value=def_date_dec, format="DD-MM-YYYY")
                date_due = st.date_input("Due Date:", value=def_due_date, format="DD-MM-YYYY")
                
            submit_btn = st.form_submit_button("💾 Save Bonus", type="primary")
            
            if submit_btn:
                if not emp_options:
                    st.error("No employees available.")
                else:
                    emp_id = emp_options[sel_emp_name]
                    if is_edit:
                        execute_query("""
                            UPDATE bonuses SET employee_id = ?, amount = ?, date_declared = ?, due_date = ? WHERE id = ?
                        """, (emp_id, amt, date_dec.strftime("%Y-%m-%d"), date_due.strftime("%Y-%m-%d"), bonus_id), commit=True)
                        st.success("Bonus updated successfully!")
                    else:
                        execute_query("""
                            INSERT INTO bonuses (employee_id, amount, date_declared, due_date) VALUES (?, ?, ?, ?)
                        """, (emp_id, amt, date_dec.strftime("%Y-%m-%d"), date_due.strftime("%Y-%m-%d")), commit=True)
                        st.success("Bonus added successfully!")
                    st.session_state.bonus_action = 'list'
                    st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
        
    else:
        # List View
        # List View                
        st.markdown("<div class='content-section'>", unsafe_allow_html=True)
        c_title, c_add, c_exp = st.columns([2, 1, 1])
        with c_title:
            st.markdown("<div class='section-title'>Recorded Bonuses</div>", unsafe_allow_html=True)
        with c_add:
            if st.button("➕ Add Bonus", type="primary", use_container_width=True):
                st.session_state.bonus_action = 'add'
                st.rerun()
        
        # Check if there is a delete confirmation pending
        if 'bonus_delete_confirm' in st.session_state and st.session_state.bonus_delete_confirm:
            del_id = st.session_state.bonus_delete_confirm
            st.warning("⚠️ Are you sure you want to delete this bonus?")
            c1, c2, c3 = st.columns([1, 1, 4])
            with c1:
                if st.button("✔️ Confirm Delete", type="primary"):
                    execute_query("DELETE FROM bonuses WHERE id = ?", (del_id,), commit=True)
                    st.session_state.bonus_delete_confirm = None
                    st.success("Bonus deleted.")
                    st.rerun()
            with c2:
                if st.button("❌ Cancel"):
                    st.session_state.bonus_delete_confirm = None
                    st.rerun()
            st.markdown("<hr/>", unsafe_allow_html=True)
            
        bonuses_data = execute_query("""
            SELECT b.id, b.employee_id, e.name as employee_name, e.role, e.date_of_joining, b.amount, b.date_declared, b.due_date
            FROM bonuses b
            JOIN employees e ON b.employee_id = e.id
            ORDER BY b.due_date DESC, b.id DESC
        """)
        
        if bonuses_data:
            df_bonuses = pd.DataFrame(bonuses_data)
            
            with c_exp:
                eur_rate = get_eur_rate()
                df_export = df_bonuses.copy()
                df_export['Role'] = df_export['role'].fillna('-')
                df_export['Date of Joining'] = df_export['date_of_joining'].apply(format_display_date)
                df_export['Amount (INR)'] = df_export['amount'].apply(format_currency_inr)
                df_export['Amount (EUR)'] = df_export['amount'].apply(lambda x: format_eur(x, eur_rate))
                df_export['Date Declared'] = df_export['date_declared'].apply(format_display_date)
                df_export['Due Date'] = df_export['due_date'].apply(format_display_date)
                df_table = df_export[[
                    'employee_id', 'employee_name', 'Role', 'Date of Joining', 
                    'Amount (INR)', 'Amount (EUR)', 'Date Declared', 'Due Date'
                ]].rename(columns={'employee_id': 'Emp ID', 'employee_name': 'Employee Name'})
                st.download_button(
                    label="📥 Export to XLSX",
                    data=to_excel(df_table),
                    file_name=f"bonuses_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            hcol0, hcol1, hcol2, hcol3, hcol4, hcol5 = st.columns([0.8, 2, 1.5, 1.5, 1.5, 1.2])
            with hcol0: st.markdown("**Emp ID**")
            with hcol1: st.markdown("**Employee Name**")
            with hcol2: st.markdown("**Amount**")
            with hcol3: st.markdown("**Date Declared**")
            with hcol4: st.markdown("**Due Date**")
            with hcol5: st.markdown("**Actions**")
            st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
            
            for idx, row in df_bonuses.iterrows():
                c0, c1, c2, c3, c4, c5 = st.columns([0.8, 2, 1.5, 1.5, 1.5, 1.2])
                with c0: st.write(row['employee_id'])
                with c1: st.write(row['employee_name'])
                with c2: st.markdown(format_currency_html(row['amount'], block=True), unsafe_allow_html=True)
                with c3: st.write(format_display_date(row['date_declared']))
                with c4: st.write(format_display_date(row['due_date']))
                with c5:
                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("✏️", key=f"edit_b_{row['id']}_{idx}", help="Edit Bonus"):
                            st.session_state.bonus_action = int(row['id'])
                            st.rerun()
                    with btn_col2:
                        if st.button("🗑️", key=f"del_b_{row['id']}_{idx}", help="Delete Bonus"):
                            st.session_state.bonus_delete_confirm = int(row['id'])
                            st.rerun()
                st.markdown("<hr style='margin: 0.2rem 0 0.45rem 0; opacity: 0.35;'>", unsafe_allow_html=True)
                
        else:
            st.info("No bonuses recorded yet.")
        st.markdown("</div>", unsafe_allow_html=True)


# --- SYSTEM SETTINGS PAGE ---
elif menu == "⚙️ System Settings":
    render_header("System Settings", "S-8")
    
    st.markdown("<div class='content-section'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Currency & Conversion Settings</div>", unsafe_allow_html=True)
    
    current_rate = get_eur_rate()
    
    col1, col2 = st.columns([1, 2])
    with col1:
        st.info(f"Current Configured Rate:\n\n**1 EUR = {current_rate} INR**")
        
    with col2:
        with st.form("settings_eur_form"):
            new_rate = st.number_input("Set new EUR to INR conversion rate:", min_value=1.0, value=float(current_rate), step=1.0)
            
            if st.form_submit_button("💾 Save Settings", type="primary"):
                execute_query("""
                    INSERT OR REPLACE INTO settings (key, value) VALUES ('eur_to_inr_rate', ?)
                """, (str(new_rate),), commit=True)
                st.success("Settings updated successfully! The EUR amounts across the application will now reflect the new rate.")
                st.rerun()
                
    st.markdown("</div>", unsafe_allow_html=True)
